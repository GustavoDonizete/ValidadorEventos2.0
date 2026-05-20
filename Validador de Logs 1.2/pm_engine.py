"""
pm_engine.py
Motor de cálculo para validação de bases de Process Mining.
6 Pilares — score final 0-100 com diagnóstico didático.

Uso via linha de comando:
    python pm_engine.py \
        --input logs.csv \
        --case_id "Numero_Processo" \
        --activity "Etapa" \
        --start_ts "Data_Inicio" \
        --end_ts "Data_Fim" \
        --output resultado.json

    Parâmetros opcionais:
        --delimiter ","          (padrão: vírgula)
        --encoding "utf-8"       (padrão: utf-8)
        --dayfirst               (flag: usa dd/mm/yyyy)
        --no_end_ts              (flag: ignora Timestamp_Fim)

Uso via API Flask (integração HTML):
    python pm_engine.py --serve --port 5000
"""

import argparse
import getpass
import json
import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

# Cache opcional — não falha se pm_cache.py não estiver presente
try:
    from pm_cache import get_default_cache as _get_cache
    _CACHE_ENABLED = True
except ImportError:
    _CACHE_ENABLED = False

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

PLACEHOLDER_VALUES = ["", " ", "-", "N/A", "NA", "NULL", "null", "None", "none", "nan", "NaN"]

PILLAR_WEIGHTS = {
    "pillar1": 20,
    "pillar2": 15,
    "pillar3": 20,
    "pillar4": 15,
    "pillar5": 15,
    "pillar6": 15,
}

# ─────────────────────────────────────────────────────────────────────────────
# I/O & NORMALIZAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

def _detect_encoding(input_path: str) -> str:
    """Detecta encoding do arquivo via BOM e fallback por tentativa."""
    try:
        with open(input_path, "rb") as f:
            raw = f.read(4)
        if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
            return "utf-16"
        if raw[:3] == b"\xef\xbb\xbf":
            return "utf-8-sig"
    except Exception:
        pass
    # Tenta encodings comuns em ordem
    for enc in ("utf-8", "cp1252", "iso-8859-1", "latin-1"):
        try:
            with open(input_path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"


def _detect_delimiter(input_path: str, encoding: str) -> str:
    """Detecta o separador do CSV lendo o cabecalho e contando ocorrencias."""
    try:
        with open(input_path, "r", encoding=encoding, errors="replace") as f:
            header = f.readline()
        counts = {
            ";":  header.count(";"),
            ",":  header.count(","),
            "\t": header.count("\t"),
            "|":  header.count("|"),
        }
        best = max(counts, key=counts.get)
        # So usa se encontrou pelo menos 1 ocorrencia
        return best if counts[best] > 0 else ","
    except Exception:
        return ","


def load_data(input_path: str, delimiter: str = "auto", encoding: str = "auto") -> pd.DataFrame:
    """
    Carrega CSV ou XLSX detectando pela extensao.
    delimiter='auto' detecta automaticamente ; , tab ou |
    encoding='auto'  detecta automaticamente via BOM e fallback
    """
    ext = Path(input_path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(input_path, dtype=str)
    elif ext in (".csv", ".tsv", ".txt"):
        # Detectar encoding
        detected_encoding = _detect_encoding(input_path) if encoding == "auto" else encoding

        # Detectar separador
        if ext == ".tsv":
            sep = "\t"
        elif delimiter == "auto":
            sep = _detect_delimiter(input_path, detected_encoding)
        else:
            sep = delimiter

        try:
            df = pd.read_csv(input_path, dtype=str, sep=sep,
                             encoding=detected_encoding, low_memory=False)
        except UnicodeDecodeError:
            df = pd.read_csv(input_path, dtype=str, sep=sep,
                             encoding="latin-1", low_memory=False)
        except Exception:
            # Ultimo recurso: python engine que e mais tolerante
            df = pd.read_csv(input_path, dtype=str, sep=sep,
                             encoding="latin-1", low_memory=False,
                             engine="python", on_bad_lines="skip")
    else:
        raise ValueError(f"Formato nao suportado: {ext}. Use CSV ou XLSX.")
    return df


def normalize_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Trim em todas as strings e substitui placeholders por NaN."""
    df = df.copy()
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()
        df[col] = df[col].replace(PLACEHOLDER_VALUES, np.nan)
    return df


def prepare_df(df: pd.DataFrame, case_col: str, act_col: str,
               start_col: str, end_col, dayfirst: bool) -> tuple:
    """
    Prepara o DataFrame para análise com otimizações de performance:

      G — Converte case_col e act_col para dtype 'category' (string repetida).
          Reduz memória ~40% e acelera groupby ~3-4x.

      B — Parseia timestamps uma única vez em __start_ts / __end_ts.
          Elimina as 12+ chamadas redundantes de pd.to_datetime nos pilares.

      F — Ordena por [case_col, __start_ts] uma única vez.
          Pilares 3, 5 e 6 reutilizam df_sorted sem re-ordenar.

    Retorna:
        df        — DataFrame com colunas __start_ts/__end_ts e dtypes otimizados
        df_sorted — df ordenado por [case_col, __start_ts]
    """
    # G: category dtype para colunas de alta repetição
    df[case_col] = df[case_col].astype("category")
    df[act_col]  = df[act_col].astype("category")

    # B: parse único de timestamps
    df["__start_ts"] = pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)
    if end_col:
        df["__end_ts"] = pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst)
    else:
        df["__end_ts"] = pd.NaT

    # F: sort único — reutilizado por P3, P5, P6
    df_sorted = df.sort_values([case_col, "__start_ts"]).reset_index(drop=True)

    return df, df_sorted


def parse_timestamps(df: pd.DataFrame, start_col: str, end_col,
                     dayfirst: bool = True) -> pd.DataFrame:
    """Parse das colunas de timestamp com coerce."""
    df = df.copy()
    df["__start_ts"] = pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)
    if end_col:
        df["__end_ts"] = pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst)
    else:
        df["__end_ts"] = pd.NaT
    return df


def validate_columns(df: pd.DataFrame, case_col: str, act_col: str,
                     start_col: str, end_col) -> None:
    """Valida colunas essenciais."""
    required = [case_col, act_col, start_col]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colunas essenciais nao encontradas no arquivo: {missing}")
    if end_col and end_col not in df.columns:
        raise ValueError(f"Coluna Timestamp_Fim '{end_col}' nao encontrada no arquivo.")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS COMPARTILHADOS
# ─────────────────────────────────────────────────────────────────────────────

def null_rate(series: pd.Series) -> float:
    return float(series.isna().sum() / max(len(series), 1))


def infer_datetime_format(series: pd.Series) -> dict:
    """
    Infere o formato de datetime dominante de uma Series de strings.

    Estrategia:
      1. Testa formatos explicitos com strptime exato (sem regex ambiguo).
      2. Agrupa por FAMILIA (ISO / BR_slash / BR_dash / US / COMPACT).
         Variacoes da mesma familia (com/sem hora, com/sem segundos)
         contam como 1 formato — evita falso-positivo quando TS_Inicio
         e TS_Fim usam o mesmo padrao mas com precisions diferentes.
      3. Em ambiguidade BR vs US, verifica dia > 12 para confirmar BR.
      4. Deriva dayfirst automaticamente do formato eleito.

    Retorna dict com:
      - fmt        : formato eleito (ex: '%d/%m/%Y %H:%M:%S')
      - dayfirst   : bool inferido
      - n_formats  : numero de FAMILIAS distintas (nao variantes)
      - parse_rate : fracao parseada com sucesso (0-1)
      - label      : descricao legivel
    """
    from datetime import datetime as _dt

    # (fmt_strptime, dayfirst, familia, label)
    CANDIDATES = [
        ("%Y-%m-%d %H:%M:%S",  False, "ISO",      "ISO datetime (yyyy-mm-dd HH:MM:SS)"),
        ("%Y-%m-%dT%H:%M:%S",  False, "ISO",      "ISO datetime com T"),
        ("%Y-%m-%d %H:%M",     False, "ISO",      "ISO datetime sem segundos"),
        ("%Y-%m-%dT%H:%M",     False, "ISO",      "ISO datetime com T sem segundos"),
        ("%Y-%m-%d",           False, "ISO",      "ISO date (yyyy-mm-dd)"),
        ("%d/%m/%Y %H:%M:%S",  True,  "BR_slash", "BR datetime (dd/mm/yyyy HH:MM:SS)"),
        ("%d/%m/%Y %H:%M",     True,  "BR_slash", "BR datetime sem segundos"),
        ("%d/%m/%Y",           True,  "BR_slash", "BR date (dd/mm/yyyy)"),
        ("%d/%m/%y",           True,  "BR_slash", "BR date 2 digitos (dd/mm/yy)"),
        ("%d-%m-%Y %H:%M:%S",  True,  "BR_dash",  "BR datetime com traco"),
        ("%d-%m-%Y %H:%M",     True,  "BR_dash",  "BR datetime com traco sem segundos"),
        ("%d-%m-%Y",           True,  "BR_dash",  "BR date com traco"),
        ("%m/%d/%Y %H:%M:%S",  False, "US",       "US datetime (mm/dd/yyyy HH:MM:SS)"),
        ("%m/%d/%Y %H:%M",     False, "US",       "US datetime sem segundos"),
        ("%m/%d/%Y",           False, "US",       "US date (mm/dd/yyyy)"),
        ("%Y%m%d%H%M%S",       False, "COMPACT",  "Compacto (yyyymmddHHMMSS)"),
        ("%Y%m%d",             False, "COMPACT",  "Compacto date (yyyymmdd)"),
    ]

    sample = series.dropna().astype(str).str.strip()

    epoch_mask   = sample.str.match(r"^\d{10,13}$")
    has_epoch    = epoch_mask.sum() > 0
    sample_clean = sample[~epoch_mask]

    if len(sample_clean) == 0:
        return {"fmt": None, "dayfirst": True, "n_formats": 1 if has_epoch else 0,
                "parse_rate": 0.0, "label": "epoch numerico" if has_epoch else "indefinido"}

    sample_vals = sample_clean.sample(min(400, len(sample_clean)), random_state=42).tolist()
    n = len(sample_vals)

    scores = {}  # fmt -> (hit_rate, dayfirst, familia, label)
    fmt_order = [c[0] for c in CANDIDATES]

    for fmt, df_flag, familia, label in CANDIDATES:
        hits = 0
        for v in sample_vals:
            # Normaliza: remove microseg, timezone, espacos
            v_clean = v.split(".")[0].split("+")[0].split("Z")[0].strip()
            try:
                _dt.strptime(v_clean, fmt)
                hits += 1
            except ValueError:
                pass
        if hits > 0:
            scores[fmt] = (hits / n, df_flag, familia, label)

    if not scores:
        parsed = pd.to_datetime(sample_clean, errors="coerce", dayfirst=True)
        rate   = parsed.notna().sum() / max(len(sample_clean), 1)
        return {"fmt": None, "dayfirst": True, "n_formats": 1,
                "parse_rate": round(rate, 3), "label": "inferido automaticamente"}

    # Eleger melhor formato: maior taxa de acerto; empate = mais especifico
    def sort_key(f):
        rate = scores[f][0]
        pos  = fmt_order.index(f) if f in fmt_order else 999
        return (rate, -pos)

    best_fmt = max(scores, key=sort_key)
    best_rate, best_dayfirst, best_familia, best_label = scores[best_fmt]

    # Contar FAMILIAS distintas com >= 20% de cobertura
    familias_sig = set(
        scores[f][2] for f in scores if scores[f][0] > 0.20
    )

    # Resolver BR_slash vs US — ambas usam dd/mm ou mm/dd com barra
    if "BR_slash" in familias_sig and "US" in familias_sig:
        try:
            first_fields = sample_clean.str.extract(r"^(\d{1,2})/").dropna()[0].astype(int)
            # Se qualquer valor tem campo > 12, so pode ser dmy (BR)
            if (first_fields > 12).any():
                familias_sig.discard("US")
            else:
                # Ambiguo: assume BR como padrao para bases brasileiras
                familias_sig.discard("US")
        except Exception:
            familias_sig.discard("US")

    n_familias = max(1, len(familias_sig)) + (1 if has_epoch else 0)

    return {
        "fmt":        best_fmt,
        "dayfirst":   best_dayfirst,
        "n_formats":  n_familias,
        "parse_rate": round(best_rate, 3),
        "label":      best_label,
    }


def detect_ts_patterns(series: pd.Series, inferred: dict = None) -> int:
    """
    Wrapper de compatibilidade — retorna n_formats do infer_datetime_format.
    Aceita resultado pre-computado via parametro 'inferred' para evitar
    rodar a inferencia duas vezes.
    """
    if inferred is not None:
        return inferred["n_formats"]
    return infer_datetime_format(series)["n_formats"]


def classify_case_id_consistency(series: pd.Series) -> tuple:
    """Classifica consistência do Case ID. Retorna (label, pts_10)."""
    clean = series.dropna().astype(str)
    if len(clean) == 0:
        return ("indefinido", 0)
    n = len(clean)
    pct_int   = clean.str.match(r"^\d+$").sum() / n
    pct_float = clean.str.match(r"^\d+\.0$").sum() / n
    pct_uuid  = clean.str.match(r"^[0-9a-fA-F\-]{32,36}$").sum() / n
    pct_alpha = clean.str.contains(r"[a-zA-Z]", na=False).sum() / n
    if pct_uuid  > 0.9: return ("uuid_consistente", 10)
    if pct_alpha > 0.9: return ("alfanumerico_consistente", 10)
    if pct_int   > 0.9: return ("numerico_consistente", 10)
    if (pct_int + pct_float) > 0.9: return ("numerico_com_variacao", 5)
    return ("alta_inconsistencia", 0)


def detect_control_chars(series: pd.Series) -> tuple:
    """Detecta caracteres de controle/encoding suspeito. Retorna (label, pts_10)."""
    ctrl_re = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f\ufffd]")
    sample = series.dropna().astype(str).head(10000)
    hits = sample.apply(lambda x: bool(ctrl_re.search(x))).sum()
    rate = hits / max(len(sample), 1)
    if rate == 0:       return ("nenhum", 10)
    if rate <= 0.01:    return ("poucos", 5)
    return ("frequentes", 0)


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 1 — Qualidade Estrutural e de Formato
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar1(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True,
                  ts_inferred: dict = None) -> dict:
    """
    Pilar 1: verifica presenca, parseabilidade e normalizacao dos campos essenciais.
    Score: 0-100. Gates de reprovacao automatica.
    ts_inferred: resultado pre-computado de infer_datetime_format (evita rodar 2x).
    Usa __start_ts / __end_ts pré-computados pelo prepare_df quando disponíveis (B).
    """
    n = len(df)
    gates_triggered = []
    diagnostics = []

    # ── Null rates ──
    nr_case  = null_rate(df[case_col])
    nr_act   = null_rate(df[act_col])
    nr_start = null_rate(df[start_col])

    # Sub 1.1 — Colunas obrigatórias (60 pts)
    def pts_null_20(rate):
        if rate == 0:         return 20
        if rate <= 0.005:     return 13
        if rate <= 0.02:      return 6
        return 0

    def pts_null_15(rate):
        if rate == 0:         return 20
        if rate <= 0.001:     return 10
        return 0

    s11_case  = pts_null_20(nr_case)
    s11_act   = pts_null_20(nr_act)
    s11_start = pts_null_15(nr_start)

    # Timestamp_Fim — penalidade se existir e tiver inválidos
    # B: usa __end_ts pré-computado se disponível
    end_penalty = 0
    nr_end_invalid = 0.0
    if end_col:
        end_parsed = df["__end_ts"] if "__end_ts" in df.columns else pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst)
        end_non_null = df[end_col].notna()
        invalid_end = end_non_null & end_parsed.isna()
        nr_end_invalid = float(invalid_end.sum() / max(n, 1))
        if nr_end_invalid == 0:       end_penalty = 0
        elif nr_end_invalid <= 0.001: end_penalty = -5
        else:                         end_penalty = -15

    score_11 = s11_case + s11_act + s11_start + end_penalty

    # Sub 1.2 — Parse de tipos e formatos (40 pts)
    # B: usa __start_ts pré-computado se disponível
    start_parsed = df["__start_ts"] if "__start_ts" in df.columns else pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)
    invalid_start_rate = float(start_parsed.isna().sum() / max(n, 1))

    ts_invalid_total = invalid_start_rate
    if end_col:
        ts_invalid_total = (ts_invalid_total + nr_end_invalid) / 2

    def pts_ts_invalid(rate):
        if rate == 0:         return 25
        if rate <= 0.001:     return 18
        if rate <= 0.01:      return 7
        return 0

    s12_ts_invalid = pts_ts_invalid(ts_invalid_total)

    all_ts = df[start_col].dropna()
    if end_col:
        all_ts = pd.concat([df[start_col].dropna(), df[end_col].dropna()])
    # Usa resultado pre-computado se disponivel, senao infere agora
    if ts_inferred is None:
        ts_inferred = infer_datetime_format(all_ts)
    n_patterns   = ts_inferred["n_formats"]
    fmt_label    = ts_inferred.get("label", "")
    s12_patterns = 15 if n_patterns == 1 else (8 if n_patterns == 2 else 0)

    # Calcular consistency_label para uso nos diagnósticos (sem pontuar aqui — movido ao Pilar 5)
    consistency_label, _s12_consistency_unused = classify_case_id_consistency(df[case_col])

    score_12 = s12_ts_invalid + s12_patterns

    # Sub 1.3 — Normalização mínima: movida para o Pilar 5
    # Calculamos apenas para métricas/diagnósticos no retorno do Pilar 1
    combined = pd.concat([df[case_col], df[act_col]]).dropna().astype(str)
    n_combined = len(combined)
    affected = combined.str.contains(r"^\s|\s$", na=False).sum()
    norm_rate = float(affected / max(n_combined, 1))
    ctrl_label, _s13_ctrl_unused = detect_control_chars(pd.concat([df[case_col], df[act_col]]))

    # ── Score final Pilar 1 (sub 1.1 + sub 1.2 = 100 pts) ──
    raw = score_11 + score_12
    score = max(0, min(100, raw))

    # ── Gates ──
    if nr_case > 0.02:
        gates_triggered.append("Case_ID nulo/vazio > 2%")
    if nr_act > 0.02:
        gates_triggered.append("Atividade nula/vazia > 2%")
    if invalid_start_rate > 0.01:
        gates_triggered.append("Timestamp_Inicio invalido > 1%")
    if ts_invalid_total > 0.01:
        gates_triggered.append("Timestamps invalidos (inicio+fim) > 1%")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos didáticos ──
    if nr_case > 0:
        diagnostics.append({
            "check": "Nulos em Case_ID",
            "valor": f"{nr_case*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if nr_case > 0.02 else "ALERTA",
            "impacto": f"-{20 - s11_case} pts",
            "descricao": f"{nr_case*100:.2f}% dos registros estao sem Case_ID. Isso impede rastrear o caminho do processo."
        })
    if nr_act > 0:
        diagnostics.append({
            "check": "Nulos em Atividade",
            "valor": f"{nr_act*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if nr_act > 0.02 else "ALERTA",
            "impacto": f"-{20 - s11_act} pts",
            "descricao": f"{nr_act*100:.2f}% dos eventos nao tem Atividade definida. Sem atividade nao ha fluxo de processo."
        })
    if invalid_start_rate > 0:
        diagnostics.append({
            "check": "Timestamp_Inicio invalido",
            "valor": f"{invalid_start_rate*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if invalid_start_rate > 0.01 else "ALERTA",
            "impacto": f"-{20 - s11_start} pts",
            "descricao": f"{invalid_start_rate*100:.2f}% dos timestamps de inicio nao puderam ser parseados. Verifique o formato de data."
        })
    if n_patterns > 1:
        diagnostics.append({
            "check": "Multiplos formatos de timestamp",
            "valor": f"{n_patterns} formatos detectados",
            "threshold": "1 formato",
            "severidade": "ALERTA",
            "impacto": f"-{15 - s12_patterns} pts",
            "descricao": (
                f"Foram encontrados {n_patterns} formatos diferentes de data na base. "
                f"Formato dominante: {fmt_label}. "
                f"Isso pode causar erros de ordenacao temporal."
            )
        })
    if end_penalty < 0:
        diagnostics.append({
            "check": "Timestamp_Fim com valores invalidos",
            "valor": f"{nr_end_invalid*100:.2f}%",
            "threshold": "0%",
            "severidade": "ALERTA",
            "impacto": f"{end_penalty} pts",
            "descricao": f"{nr_end_invalid*100:.2f}% dos timestamps de fim sao invalidos. Calculos de duracao ficam comprometidos."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "1.1_presenca_colunas": {"score": max(0, score_11), "max": 60},
            "1.2_parse_formatos":   {"score": score_12,         "max": 40},
        },
        "metrics": {
            "null_rate_case_id":          round(nr_case * 100, 3),
            "null_rate_atividade":         round(nr_act * 100, 3),
            "null_rate_timestamp_inicio":  round(nr_start * 100, 3),
            "invalid_rate_timestamp":      round(ts_invalid_total * 100, 3),
            "n_ts_patterns":               n_patterns,
            "ts_format_label":             fmt_label,
            "ts_parse_rate":               ts_inferred.get("parse_rate", 1.0),
            "dayfirst_used":               dayfirst,
            "case_id_consistency":         consistency_label,
            "control_chars":               ctrl_label,
        },
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 2 — Completude & Cobertura
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar2(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True) -> dict:
    """
    Pilar 2: volume, densidade temporal e cobertura por case.
    """
    n = len(df)
    gates_triggered = []
    diagnostics = []

    # ── Sub 2.1 Completude por coluna-chave (30 pts) ──
    # nr_case foi movido para Pilar 6; aqui avaliamos apenas act, start e end
    nr_case  = null_rate(df[case_col])   # mantido para gates
    nr_act   = null_rate(df[act_col])
    nr_start = null_rate(df[start_col])

    def pts_act(rate):
        if rate == 0:       return 10
        if rate <= 0.005:   return 7
        if rate <= 0.02:    return 3
        return 0

    def pts_start(rate):
        if rate == 0:       return 20
        if rate <= 0.001:   return 13
        if rate <= 0.01:    return 6
        return 0

    s21 = pts_act(nr_act) + pts_start(nr_start)

    if end_col:
        nr_end = null_rate(df[end_col])
        s21_end = 5 if nr_end <= 0.01 else (3 if nr_end <= 0.05 else 0)
        s21 += s21_end

    # ── Sub 2.2 Densidade por case (50 pts) ──
    valid = df[[case_col, act_col]].copy()
    valid = valid[valid[case_col].notna() & valid[act_col].notna()]
    events_per_case = valid.groupby(case_col, sort=False, observed=True).size()
    total_cases = len(events_per_case)
    total_events = int(events_per_case.sum())
    pct_lt2 = float((events_per_case < 2).sum() / max(total_cases, 1))
    median_events = float(events_per_case.median()) if total_cases > 0 else 0

    def pts_pct_lt2(rate):
        if rate <= 0.05:    return 35
        if rate <= 0.15:    return 26
        if rate <= 0.30:    return 14
        if rate <= 0.60:    return 6
        return 0

    def pts_median(m):
        if m >= 6:  return 15
        if m >= 4:  return 12
        if m == 3:  return 8
        if m == 2:  return 4
        return 0

    s22 = pts_pct_lt2(pct_lt2) + pts_median(median_events)

    # ── Sub 2.3 Cobertura temporal (20 pts — apenas período total) ──
    # pts_gap movido para Pilar 3; pts_density removido
    # B: usa __start_ts / __end_ts pré-computados se disponíveis
    start_parsed = df["__start_ts"] if "__start_ts" in df.columns else pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)
    max_ts = start_parsed
    if end_col:
        end_parsed = df["__end_ts"] if "__end_ts" in df.columns else pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst)
        max_ts = start_parsed.combine(end_parsed, lambda a, b: max(a, b) if pd.notna(a) and pd.notna(b) else (a if pd.notna(a) else b))

    min_ts_val = start_parsed.min()
    max_ts_val = max_ts.max()

    period_days = 0
    if pd.notna(min_ts_val) and pd.notna(max_ts_val):
        period_days = (max_ts_val - min_ts_val).days

    def pts_period(d):
        if d >= 90: return 20
        if d >= 30: return 14
        if d >= 7:  return 7
        return 0

    # Maior gap — calculado para métricas e passado para o Pilar 3
    active_dates = start_parsed.dropna().dt.date
    unique_dates = sorted(active_dates.unique())
    gap_max = 0
    if len(unique_dates) > 1:
        gaps = [(unique_dates[i+1] - unique_dates[i]).days for i in range(len(unique_dates)-1)]
        gap_max = max(gaps)

    events_per_day  = total_events / max(period_days, 1)
    events_per_week = events_per_day * 7

    s23 = pts_period(period_days)
    score = min(100, s21 + s22 + s23)

    # ── Gates ──
    if pct_lt2 > 0.90:
        gates_triggered.append(f"Cases com < 2 eventos > 90% ({pct_lt2*100:.1f}%)")
    if median_events < 2:
        gates_triggered.append(f"Mediana de eventos por case < 2 ({median_events:.1f})")
    if period_days < 2:
        gates_triggered.append(f"Periodo total < 2 dias ({period_days} dias)")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos ──
    if pct_lt2 > 0.05:
        diagnostics.append({
            "check": "Cases com poucos eventos",
            "valor": f"{pct_lt2*100:.1f}%",
            "threshold": "<=5%",
            "severidade": "CRITICO" if pct_lt2 > 0.90 else "ALERTA",
            "impacto": f"-{35 - pts_pct_lt2(pct_lt2)} pts",
            "descricao": f"{pct_lt2*100:.1f}% dos cases tem menos de 2 eventos. Um processo com 1 evento nao tem fluxo mensuravel."
        })
    if median_events < 4:
        diagnostics.append({
            "check": "Mediana baixa de eventos por case",
            "valor": f"{median_events:.1f} eventos",
            "threshold": ">=6",
            "severidade": "CRITICO" if median_events < 2 else "ALERTA",
            "impacto": f"-{15 - pts_median(median_events)} pts",
            "descricao": f"A mediana de eventos por case e {median_events:.1f}. Para Process Mining significativo, o ideal e >= 6 eventos por processo."
        })
    if period_days < 30:
        diagnostics.append({
            "check": "Periodo temporal curto",
            "valor": f"{period_days} dias",
            "threshold": ">=90 dias",
            "severidade": "ALERTA",
            "impacto": f"-{20 - pts_period(period_days)} pts",
            "descricao": f"A base cobre apenas {period_days} dias. Analises de sazonalidade e tendencia ficam limitadas."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "2.1_completude_colunas": {"score": s21, "max": 30 if not end_col else 35},
            "2.2_densidade_por_case": {"score": s22, "max": 50},
            "2.3_cobertura_temporal": {"score": s23, "max": 20},
        },
        "metrics": {
            "total_cases":            total_cases,
            "total_events":           total_events,
            "pct_cases_lt2":          round(pct_lt2 * 100, 2),
            "median_events_per_case": round(median_events, 1),
            "period_days":            period_days,
            "gap_max_days":           gap_max,
            "events_per_week":        round(events_per_week, 1),
        },
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 3 — Integridade Temporal
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar3(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True,
                  df_sorted: "pd.DataFrame | None" = None) -> dict:
    """
    Pilar 3: coerência lógica da sequência temporal dos eventos.
    df_sorted: DataFrame pré-ordenado por [case_col, __start_ts] (otimização F).
               Se None, ordena internamente.
    Usa __start_ts / __end_ts pré-computados (otimização B).
    """
    gates_triggered = []
    diagnostics = []

    # B+F: usar df_sorted pré-computado quando disponível
    _base = df_sorted if df_sorted is not None else df
    has_precomputed = "__start_ts" in _base.columns

    cols = [case_col, act_col]
    if has_precomputed:
        cols += ["__start_ts"]
        if end_col: cols += ["__end_ts"]
    else:
        cols += [start_col]
        if end_col: cols += [end_col]

    valid = _base[[c for c in cols if c in _base.columns]].copy()
    valid = valid[valid[case_col].notna() & valid[act_col].notna()]

    if has_precomputed:
        valid["__start"] = valid["__start_ts"]
        if end_col and "__end_ts" in valid.columns:
            valid["__end"] = valid["__end_ts"]
    else:
        valid["__start"] = pd.to_datetime(valid[start_col], errors="coerce", dayfirst=dayfirst)
        if end_col:
            valid["__end"] = pd.to_datetime(valid[end_col], errors="coerce", dayfirst=dayfirst)

    n_valid = len(valid)

    # ── Sub 3.1 Inversões de ordem temporal por case (30 pts) ──
    # F: já está ordenado se df_sorted foi passado
    valid_ts = valid[valid["__start"].notna()].copy()
    if df_sorted is None:
        valid_ts = valid_ts.sort_values([case_col, "__start"])
    valid_ts["__prev_start"] = valid_ts.groupby(case_col, sort=False, observed=True)["__start"].shift(1)
    inversions = (valid_ts["__start"] < valid_ts["__prev_start"]).sum()
    inversion_rate = float(inversions / max(n_valid, 1))

    def pts_inversion(rate):
        if rate == 0:       return 30
        if rate <= 0.001:   return 22
        if rate <= 0.01:    return 11
        if rate <= 0.05:    return 4
        return 0

    s31 = pts_inversion(inversion_rate)

    # ── Sub 3.2 Durações negativas (25 pts — só se end_col existir) ──
    s32 = 25
    neg_duration_rate = 0.0
    zero_duration_rate = 0.0
    if end_col:
        has_both = valid["__start"].notna() & valid["__end"].notna()
        paired = valid[has_both].copy()
        if len(paired) > 0:
            duration = (paired["__end"] - paired["__start"]).dt.total_seconds()
            neg_duration_rate  = float((duration < 0).sum() / max(len(paired), 1))
            zero_duration_rate = float((duration == 0).sum() / max(len(paired), 1))

            def pts_neg(rate):
                if rate == 0:       return 25
                if rate <= 0.001:   return 17
                if rate <= 0.01:    return 7
                return 0

            s32 = pts_neg(neg_duration_rate)
    else:
        s32 = 25  # sem end_col, subcategoria não se aplica — pontuação máxima

    # ── Sub 3.3 Gaps absurdos dentro do mesmo case (25 pts) ──
    # F: reutiliza valid_ts já ordenado (ou ordena se não havia df_sorted)
    valid_ts2 = valid[valid["__start"].notna()].copy()
    if df_sorted is None:
        valid_ts2 = valid_ts2.sort_values([case_col, "__start"])
    valid_ts2["__next_start"] = valid_ts2.groupby(case_col, sort=False, observed=True)["__start"].shift(-1)
    valid_ts2["__gap_days"] = (valid_ts2["__next_start"] - valid_ts2["__start"]).dt.days
    intra_gaps = valid_ts2["__gap_days"].dropna()
    absurd_threshold = 365
    absurd_rate = float((intra_gaps > absurd_threshold).sum() / max(len(intra_gaps), 1))

    def pts_absurd(rate):
        if rate == 0:       return 25
        if rate <= 0.001:   return 17
        if rate <= 0.01:    return 8
        return 0

    s33 = pts_absurd(absurd_rate)

    # ── Sub 3.4 Gap máximo entre dias com eventos (20 pts — vindo do Pilar 2) ──
    # B: usa __start_ts pré-computado se disponível
    _start_p3 = df["__start_ts"] if "__start_ts" in df.columns else pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)
    active_dates_p3 = _start_p3.dropna().dt.date
    unique_dates_p3 = sorted(active_dates_p3.unique())
    gap_max = 0
    if len(unique_dates_p3) > 1:
        gaps_p3 = [(unique_dates_p3[i+1] - unique_dates_p3[i]).days for i in range(len(unique_dates_p3)-1)]
        gap_max = max(gaps_p3)

    def pts_gap(g):
        if g <= 3:  return 20
        if g <= 7:  return 14
        if g <= 30: return 7
        return 0

    s34 = pts_gap(gap_max)

    raw = s31 + s32 + s33 + s34
    score = min(100, max(0, raw))

    # ── Gates ──
    if inversion_rate > 0.05:
        gates_triggered.append(f"Inversoes de ordem temporal > 5% ({inversion_rate*100:.1f}%)")
    if neg_duration_rate > 0.01:
        gates_triggered.append(f"Duracoes negativas > 1% ({neg_duration_rate*100:.1f}%)")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos ──
    if inversion_rate > 0:
        diagnostics.append({
            "check": "Inversoes de ordem temporal",
            "valor": f"{inversion_rate*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if inversion_rate > 0.05 else "ALERTA",
            "impacto": f"-{30 - s31} pts",
            "descricao": f"{inversion_rate*100:.2f}% dos eventos estao fora de ordem cronologica dentro do mesmo case. Isso distorce o mapa de processo e calculos de tempo de ciclo."
        })
    if neg_duration_rate > 0:
        diagnostics.append({
            "check": "Duracoes negativas (fim antes do inicio)",
            "valor": f"{neg_duration_rate*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if neg_duration_rate > 0.01 else "ALERTA",
            "impacto": f"-{25 - s32} pts",
            "descricao": f"{neg_duration_rate*100:.2f}% dos eventos tem Timestamp_Fim anterior ao Timestamp_Inicio. Erro de registro ou inversao de colunas."
        })
    if zero_duration_rate > 0.1:
        diagnostics.append({
            "check": "Duracoes zeradas",
            "valor": f"{zero_duration_rate*100:.1f}%",
            "threshold": "<=10%",
            "severidade": "INFO",
            "impacto": "-0 pts",
            "descricao": f"{zero_duration_rate*100:.1f}% dos eventos tem duracao zero. Pode ser atividade instantanea ou dado sintetico."
        })
    if absurd_rate > 0:
        diagnostics.append({
            "check": "Gaps absurdos entre eventos do mesmo case",
            "valor": f"{absurd_rate*100:.3f}%",
            "threshold": "0%",
            "severidade": "ALERTA",
            "impacto": f"-{25 - s33} pts",
            "descricao": f"Existem pares de eventos do mesmo case com mais de {absurd_threshold} dias de diferenca. Pode indicar mistura de periodos distintos."
        })
    if gap_max > 7:
        diagnostics.append({
            "check": "Gap sem eventos na base",
            "valor": f"{gap_max} dias consecutivos",
            "threshold": "<=3 dias",
            "severidade": "ALERTA",
            "impacto": f"-{20 - s34} pts",
            "descricao": f"Existe um periodo de {gap_max} dias consecutivos sem nenhum evento. Pode indicar dados incompletos ou sistema fora do ar."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "3.1_inversao_ordem":     {"score": s31, "max": 30},
            "3.2_duracoes_negativas": {"score": s32, "max": 25},
            "3.3_gaps_absurdos":      {"score": s33, "max": 25},
            "3.4_gap_max_eventos":    {"score": s34, "max": 20},
        },
        "metrics": {
            "inversion_rate_pct":     round(inversion_rate * 100, 3),
            "neg_duration_rate_pct":  round(neg_duration_rate * 100, 3),
            "zero_duration_rate_pct": round(zero_duration_rate * 100, 3),
            "absurd_gap_rate_pct":    round(absurd_rate * 100, 3),
            "absurd_threshold_days":  absurd_threshold,
            "gap_max_days":           gap_max,
        },
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 4 — Unicidade & Duplicidade
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar4(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True) -> dict:
    """
    Pilar 4: detecta duplicatas exatas e colisões de eventos.
    """
    gates_triggered = []
    diagnostics = []

    n = len(df)
    # B: usa __start_ts / __end_ts pré-computados se disponíveis
    start_parsed = df["__start_ts"] if "__start_ts" in df.columns else pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst)

    # Chave de evento
    if end_col:
        end_parsed = df["__end_ts"] if "__end_ts" in df.columns else pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst)
        event_key = pd.DataFrame({
            "case": df[case_col],
            "act":  df[act_col],
            "start": start_parsed,
            "end":   end_parsed,
        })
    else:
        event_key = pd.DataFrame({
            "case":  df[case_col],
            "act":   df[act_col],
            "start": start_parsed,
        })

    # ── Sub 4.1 Duplicatas exatas (60 pts) ──
    n_unique = len(event_key.drop_duplicates())
    dup_rows = n - n_unique
    dup_rate = float(dup_rows / max(n, 1))

    def pts_dup(rate):
        if rate == 0:           return 60
        if rate <= 0.0005:      return 50
        if rate <= 0.002:       return 35
        if rate <= 0.01:        return 15
        return 0

    s41 = pts_dup(dup_rate)

    # ── Sub 4.2 Colisões (40 pts) ──
    collision_key = pd.DataFrame({
        "case":  df[case_col],
        "act":   df[act_col],
        "start": start_parsed,
    })
    collision_key_valid = collision_key[collision_key["case"].notna()].copy()
    grp = collision_key_valid.groupby(["case", "act", "start"], sort=False, observed=True).size()
    colliding_keys = grp[grp > 1].reset_index()
    cases_with_collision = colliding_keys["case"].nunique() if len(colliding_keys) > 0 else 0
    total_cases = df[case_col].nunique()
    collision_case_rate = float(cases_with_collision / max(total_cases, 1))

    def pts_collision(rate):
        if rate <= 0.01:    return 40
        if rate <= 0.05:    return 30
        if rate <= 0.15:    return 18
        if rate <= 0.30:    return 8
        return 0

    s42 = pts_collision(collision_case_rate)
    score = min(100, s41 + s42)

    # ── Top offenders por atividade ──
    dup_mask = event_key.duplicated(keep=False)
    dup_by_act = {}
    if dup_rows > 0 and act_col in df.columns:
        dup_counts = df[dup_mask][act_col].value_counts().head(10)
        dup_by_act = dup_counts.to_dict()

    # ── Gates ──
    if dup_rate > 0.01:
        gates_triggered.append(f"Taxa de duplicatas exatas > 1% ({dup_rate*100:.2f}%)")
    if collision_case_rate > 0.30:
        gates_triggered.append(f"Cases com colisao > 30% ({collision_case_rate*100:.1f}%)")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos ──
    if dup_rows > 0:
        diagnostics.append({
            "check": "Duplicatas exatas",
            "valor": f"{dup_rate*100:.3f}% ({dup_rows} linhas)",
            "threshold": "0%",
            "severidade": "CRITICO" if dup_rate > 0.01 else "ALERTA",
            "impacto": f"-{60 - s41} pts",
            "descricao": f"{dup_rows} eventos duplicados detectados ({dup_rate*100:.3f}%). Duplicatas inflam frequencias de atividades e distorcem calculos de loops no mapa de processo."
        })
    if cases_with_collision > 0:
        diagnostics.append({
            "check": "Colisoes de eventos (ambiguidade)",
            "valor": f"{collision_case_rate*100:.1f}% dos cases",
            "threshold": "<=1%",
            "severidade": "CRITICO" if collision_case_rate > 0.30 else "ALERTA",
            "impacto": f"-{40 - s42} pts",
            "descricao": f"{cases_with_collision} cases tem eventos com mesmo Case_ID, Atividade e Timestamp_Inicio mas fins diferentes. Pode ser reprocessamento ou erro de carga."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "4.1_duplicatas_exatas": {"score": s41, "max": 60},
            "4.2_colisoes":          {"score": s42, "max": 40},
        },
        "metrics": {
            "total_rows":              n,
            "total_cases":             total_cases,
            "dup_rows":                dup_rows,
            "dup_rate_pct":            round(dup_rate * 100, 4),
            "cases_with_collision":    cases_with_collision,
            "collision_case_rate_pct": round(collision_case_rate * 100, 2),
        },
        "top_offenders_by_activity": dup_by_act,
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 5 — Coerência do Case ID
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar5(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True,
                  df_sorted: "pd.DataFrame | None" = None) -> dict:
    """
    Pilar 5: detecta fragmentação e mescla de Case IDs via análise de percentis.
    df_sorted: DataFrame pré-ordenado por [case_col, __start_ts] (otimização F).
    Usa __start_ts / __end_ts pré-computados (otimização B).
    """
    gates_triggered = []
    diagnostics = []
    EPS = 1e-9

    # B+F: usar df_sorted pré-computado quando disponível
    _base = df_sorted if df_sorted is not None else df
    has_precomputed = "__start_ts" in _base.columns

    cols = [case_col, act_col]
    if has_precomputed:
        cols += ["__start_ts"]
        if end_col: cols += ["__end_ts"]
    else:
        cols += [start_col]
        if end_col: cols += [end_col]

    valid = _base[[c for c in cols if c in _base.columns]].copy()
    valid = valid[valid[case_col].notna() & valid[act_col].notna()]

    if has_precomputed:
        valid["__start"] = valid["__start_ts"]
        if end_col and "__end_ts" in valid.columns:
            valid["__end"] = valid["__end_ts"]
    else:
        valid["__start"] = pd.to_datetime(valid[start_col], errors="coerce", dayfirst=dayfirst)
        if end_col:
            valid["__end"] = pd.to_datetime(valid[end_col], errors="coerce", dayfirst=dayfirst)

    grp = valid[valid["__start"].notna()].groupby(case_col, sort=False, observed=True)

    events_per_case = grp["__start"].count()
    distinct_acts   = grp[act_col].nunique()

    # Duração por case
    if end_col:
        case_max_end   = grp["__end"].max()
        case_min_start = grp["__start"].min()
        duration_sec   = (case_max_end - case_min_start).dt.total_seconds().fillna(0)
    else:
        case_max_start = grp["__start"].max()
        case_min_start = grp["__start"].min()
        duration_sec   = (case_max_start - case_min_start).dt.total_seconds().fillna(0)

    # Max repeat dentro do case
    max_repeat = grp[act_col].apply(lambda x: x.value_counts().max() if len(x) > 0 else 1)

    n_cases      = len(events_per_case)
    total_events = int(events_per_case.sum())

    p95_events   = float(np.percentile(events_per_case, 95))
    p99_events   = float(np.percentile(events_per_case, 99))
    p95_acts     = float(np.percentile(distinct_acts, 95))
    p99_acts     = float(np.percentile(distinct_acts, 99))
    p95_dur      = float(np.percentile(duration_sec, 95))
    p99_dur      = float(np.percentile(duration_sec, 99))
    p99_repeat   = float(np.percentile(max_repeat, 99))

    r_events   = p99_events  / max(p95_events, EPS)
    r_acts     = p99_acts    / max(p95_acts, EPS)
    r_duration = p99_dur     / max(p95_dur, EPS)

    # Concentração top 1%
    top_k = max(1, math.ceil(0.01 * n_cases))
    top_cases_events = events_per_case.nlargest(top_k).sum()
    top1pct_share = float(top_cases_events / max(total_events, 1))

    # ── Sub 5.1 Fragmentação (35 pts) via r_duration ──
    def pts_rdur(r):
        if r <= 2:   return 35
        if r <= 5:   return 25
        if r <= 10:  return 10
        return 0

    s51 = pts_rdur(r_duration)

    # ── Sub 5.2 Mescla (35 pts) ──
    def pts_revents(r):
        if r <= 2:   return 14
        if r <= 5:   return 10
        if r <= 10:  return 4
        return 0

    def pts_racts(r):
        if r <= 2:   return 14
        if r <= 4:   return 10
        if r <= 8:   return 4
        return 0

    def pts_top1(share):
        if share <= 0.20: return 7
        if share <= 0.35: return 4
        if share <= 0.50: return 2
        return 0

    s52 = pts_revents(r_events) + pts_racts(r_acts) + pts_top1(top1pct_share)

    # ── Sub 5.3 Normalização & Consistência do Case ID (30 pts — vindo do Pilar 1) ──
    consistency_label_p5, s53_consistency = classify_case_id_consistency(df[case_col])

    combined_p5 = pd.concat([df[case_col], df[act_col]]).dropna().astype(str)
    n_combined_p5 = len(combined_p5)
    affected_p5   = combined_p5.str.contains(r"^\s|\s$", na=False).sum()
    norm_rate_p5  = float(affected_p5 / max(n_combined_p5, 1))
    s53_norm      = 10 if norm_rate_p5 <= 0.005 else (5 if norm_rate_p5 <= 0.02 else 0)

    ctrl_label_p5, s53_ctrl = detect_control_chars(pd.concat([df[case_col], df[act_col]]))

    s53 = s53_consistency + s53_norm + s53_ctrl

    score = min(100, s51 + s52 + s53)

    # Loop severity
    loop_severity = "INFO" if p99_repeat <= 10 else ("WARN" if p99_repeat <= 50 else "WARN_FORTE")

    # ── Gates ──
    if top1pct_share > 0.60:
        gates_triggered.append(f"Top 1% cases concentra > 60% eventos ({top1pct_share*100:.1f}%)")
    if r_events > 20 and r_acts > 10:
        gates_triggered.append(f"Mescla severa: r_events={r_events:.1f} e r_acts={r_acts:.1f}")
    if r_duration > 20:
        gates_triggered.append(f"Fragmentacao severa: r_duration={r_duration:.1f}")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos ──
    if r_duration > 5:
        diagnostics.append({
            "check": "Fragmentacao de Case ID",
            "valor": f"r_duration={r_duration:.1f}",
            "threshold": "<=2",
            "severidade": "CRITICO" if r_duration > 10 else "ALERTA",
            "impacto": f"-{35 - s51} pts",
            "descricao": f"A razao P99/P95 de duracao dos cases e {r_duration:.1f}x. Isso sugere que parte dos cases esta fragmentada — um processo real pode estar dividido em multiplos Case IDs curtos."
        })
    if r_events > 5:
        diagnostics.append({
            "check": "Mescla de Case ID (volume de eventos)",
            "valor": f"r_events={r_events:.1f}",
            "threshold": "<=2",
            "severidade": "CRITICO" if r_events > 10 else "ALERTA",
            "impacto": f"-{14 - min(pts_revents(r_events), 14)} pts",
            "descricao": f"A razao P99/P95 de eventos por case e {r_events:.1f}x. Alguns cases tem volume desproporcional, indicando possivel mescla de processos distintos num unico Case ID."
        })
    if top1pct_share > 0.35:
        diagnostics.append({
            "check": "Concentracao no top 1% de cases",
            "valor": f"{top1pct_share*100:.1f}%",
            "threshold": "<=20%",
            "severidade": "CRITICO" if top1pct_share > 0.60 else "ALERTA",
            "impacto": f"-{7 - pts_top1(top1pct_share)} pts",
            "descricao": f"O top 1% dos cases concentra {top1pct_share*100:.1f}% de todos os eventos. Poucos cases 'monster' dominam a base, o que pode gerar um mapa de processo distorcido."
        })
    if loop_severity != "INFO":
        diagnostics.append({
            "check": "Loops anomalos (repeticao de atividade)",
            "valor": f"P99 de max repeticoes = {p99_repeat:.0f}x",
            "threshold": "<=10",
            "severidade": loop_severity,
            "impacto": "-0 pts",
            "descricao": f"Em 1% dos cases, uma mesma atividade se repete mais de {p99_repeat:.0f} vezes. Pode ser retrabalho excessivo ou dado de sistema que cria eventos automaticamente."
        })
    if s53_consistency < 10:
        diagnostics.append({
            "check": "Inconsistencia no Case_ID",
            "valor": consistency_label_p5,
            "threshold": "consistente",
            "severidade": "ALERTA",
            "impacto": f"-{10 - s53_consistency} pts",
            "descricao": "O Case_ID mistura formatos (ex: numerico e alfanumerico). Isso pode indicar fusao de bases distintas."
        })
    if s53_norm < 10:
        diagnostics.append({
            "check": "Valores com espacos nao normalizados",
            "valor": f"{norm_rate_p5*100:.2f}%",
            "threshold": "<=0.5%",
            "severidade": "ALERTA",
            "impacto": f"-{10 - s53_norm} pts",
            "descricao": f"{norm_rate_p5*100:.2f}% dos valores de Case_ID ou Atividade possuem espacos extras (leading/trailing). Isso pode criar IDs duplicados logicos."
        })
    if s53_ctrl < 10:
        diagnostics.append({
            "check": "Caracteres de controle detectados",
            "valor": ctrl_label_p5,
            "threshold": "nenhum",
            "severidade": "ALERTA",
            "impacto": f"-{10 - s53_ctrl} pts",
            "descricao": "Foram detectados caracteres de controle ou encoding suspeito em Case_ID ou Atividade. Podem causar problemas silenciosos de comparacao de strings."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "5.1_fragmentacao":         {"score": s51, "max": 35},
            "5.2_mescla":               {"score": s52, "max": 35},
            "5.3_normalizacao_case_id": {"score": s53, "max": 30},
        },
        "metrics": {
            "n_cases":              n_cases,
            "total_events":         total_events,
            "p95_events":           round(p95_events, 1),
            "p99_events":           round(p99_events, 1),
            "r_events":             round(r_events, 2),
            "p95_distinct_acts":    round(p95_acts, 1),
            "p99_distinct_acts":    round(p99_acts, 1),
            "r_acts":               round(r_acts, 2),
            "p95_duration_h":       round(p95_dur / 3600, 2),
            "p99_duration_h":       round(p99_dur / 3600, 2),
            "r_duration":           round(r_duration, 2),
            "top1pct_event_share":  round(top1pct_share * 100, 2),
            "p99_max_repeat":       round(p99_repeat, 1),
            "loop_severity":        loop_severity,
            "case_id_consistency":  consistency_label_p5,
            "norm_rate_pct":        round(norm_rate_p5 * 100, 3),
            "control_chars":        ctrl_label_p5,
        },
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PILAR 6 — Minerabilidade & Qualidade de Atividade
# ─────────────────────────────────────────────────────────────────────────────

def score_pillar6(df: pd.DataFrame, case_col: str, act_col: str,
                  start_col: str, end_col, dayfirst: bool = True,
                  df_sorted: "pd.DataFrame | None" = None) -> dict:
    """
    Pilar 6: vocabulário de atividades e diversidade de variantes de processo.
    df_sorted: DataFrame pré-ordenado por [case_col, __start_ts] (otimização F).
    Usa __start_ts / __end_ts pré-computados (otimização B).
    """
    gates_triggered = []
    diagnostics = []

    # B+F: usar df_sorted pré-computado quando disponível
    _base = df_sorted if df_sorted is not None else df
    has_precomputed = "__start_ts" in _base.columns

    cols = [case_col, act_col]
    if has_precomputed:
        cols += ["__start_ts"]
        if end_col: cols += ["__end_ts"]
    else:
        cols += [start_col]
        if end_col: cols += [end_col]

    valid = _base[[c for c in cols if c in _base.columns]].copy()
    valid = valid[valid[case_col].notna() & valid[act_col].notna()]

    if has_precomputed:
        valid["__start"] = valid["__start_ts"]
        if end_col and "__end_ts" in valid.columns:
            valid["__end"] = valid["__end_ts"]
    else:
        valid["__start"] = pd.to_datetime(valid[start_col], errors="coerce", dayfirst=dayfirst)
        if end_col:
            valid["__end"] = pd.to_datetime(valid[end_col], errors="coerce", dayfirst=dayfirst)

    total_cases  = valid[case_col].nunique()
    total_events = len(valid)

    # ── Sub 6.1 Qualidade do vocabulário (30 pts) ──
    act_counts   = valid[act_col].value_counts()
    n_activities = len(act_counts)
    rare_acts    = (act_counts == 1).sum()
    rare_rate    = float(rare_acts / max(n_activities, 1))

    def pts_cardinality(n):
        if 5 <= n <= 200:   return 20
        if 2 <= n <= 4 or 201 <= n <= 500: return 12
        if n == 1 or 501 <= n <= 1000: return 5
        return 0

    def pts_rare_acts(rate):
        if rate <= 0.10:    return 10
        if rate <= 0.30:    return 7
        if rate <= 0.60:    return 3
        return 0

    s61 = pts_cardinality(n_activities) + pts_rare_acts(rare_rate)

    # ── Sub 6.2 Diversidade de variantes (40 pts) ──
    # F: df_sorted já está ordenado — evita re-sort
    sorted_valid = valid[valid["__start"].notna()]
    if df_sorted is None:
        sorted_valid = sorted_valid.sort_values([case_col, "__start"])
    variants = sorted_valid.groupby(case_col, sort=False, observed=True)[act_col].apply(tuple)
    variant_counts = variants.value_counts()
    n_variants     = len(variant_counts)
    variant_ratio  = float(n_variants / max(total_cases, 1))

    top10_cases    = variant_counts.head(10).sum()
    top10_coverage = float(top10_cases / max(total_cases, 1))

    rare_variant_cases  = variant_counts[variant_counts == 1].sum()
    rare_variants_rate  = float(rare_variant_cases / max(total_cases, 1))

    def pts_variant_ratio(r):
        if r <= 0.1:    return 13
        if r <= 0.3:    return 9
        if r <= 0.6:    return 4
        return 0

    def pts_top10(cov):
        if cov >= 0.80: return 17
        if cov >= 0.60: return 12
        if cov >= 0.40: return 5
        return 0

    def pts_rare_variants(rate):
        if rate <= 0.30: return 10
        if rate <= 0.60: return 6
        if rate <= 0.80: return 3
        return 0

    s62 = pts_variant_ratio(variant_ratio) + pts_top10(top10_coverage) + pts_rare_variants(rare_variants_rate)

    # ── Sub 6.3 Sobreposição de eventos (20 pts — só se end_col) ──
    s63 = None
    overlap_case_rate = None
    if end_col:
        valid_end = valid[valid["__start"].notna() & valid["__end"].notna()].copy()
        if df_sorted is None:
            valid_end = valid_end.sort_values([case_col, "__start"])
        valid_end["__prev_end"] = valid_end.groupby(case_col, sort=False, observed=True)["__end"].shift(1)
        overlap_mask = valid_end["__start"] < valid_end["__prev_end"]
        cases_with_overlap = valid_end[overlap_mask][case_col].nunique()
        overlap_case_rate  = float(cases_with_overlap / max(total_cases, 1))

        def pts_overlap(rate):
            if rate <= 0.10: return 20
            if rate <= 0.30: return 14
            if rate <= 0.50: return 6
            return 0

        s63 = pts_overlap(overlap_case_rate)

    # ── Sub 6.4 Completude do Case ID (10 pts — vindo do Pilar 2) ──
    nr_case_p6 = null_rate(df[case_col])

    def pts_case_null(rate):
        if rate == 0:       return 10
        if rate <= 0.005:   return 7
        if rate <= 0.02:    return 3
        return 0

    s64 = pts_case_null(nr_case_p6)

    # ── Score final ──
    if s63 is not None:
        raw_score = s61 + s62 + s63 + s64
        score = min(100, raw_score)
    else:
        raw_score = s61 + s62 + s64
        score = min(100, int((raw_score / 80) * 100))

    # ── Gates ──
    if variant_ratio > 0.9 and top10_coverage < 0.2:
        gates_triggered.append(f"Cada case e quase unico (variant_ratio={variant_ratio:.2f}, top10={top10_coverage*100:.1f}%)")
    if n_activities > 5000:
        gates_triggered.append(f"Mais de 5000 atividades distintas ({n_activities}) — log tecnico, nao de processo")

    status = "INAPTO" if gates_triggered else "APTO"

    # ── Diagnósticos ──
    if n_activities < 5:
        diagnostics.append({
            "check": "Poucas atividades distintas",
            "valor": f"{n_activities} atividades",
            "threshold": "5-200",
            "severidade": "ALERTA",
            "impacto": f"-{20 - pts_cardinality(n_activities)} pts",
            "descricao": f"A base tem apenas {n_activities} atividades distintas. Granularidade muito baixa — o mapa de processo sera superficial e pouco informativo."
        })
    elif n_activities > 500:
        diagnostics.append({
            "check": "Excesso de atividades distintas",
            "valor": f"{n_activities} atividades",
            "threshold": "5-200",
            "severidade": "CRITICO" if n_activities > 5000 else "ALERTA",
            "impacto": f"-{20 - pts_cardinality(n_activities)} pts",
            "descricao": f"A base tem {n_activities} atividades distintas. Pode ser um log tecnico nao padronizado. Recomenda-se agrupar ou filtrar atividades antes da analise."
        })
    if rare_rate > 0.30:
        diagnostics.append({
            "check": "Alta taxa de atividades raras",
            "valor": f"{rare_rate*100:.1f}%",
            "threshold": "<=10%",
            "severidade": "ALERTA",
            "impacto": f"-{10 - pts_rare_acts(rare_rate)} pts",
            "descricao": f"{rare_rate*100:.1f}% das atividades aparecem apenas 1 vez. Indica falta de padronizacao nos nomes de atividade ou dados de baixa qualidade."
        })
    if variant_ratio > 0.5:
        diagnostics.append({
            "check": "Muitas variantes de processo",
            "valor": f"{n_variants} variantes ({variant_ratio*100:.1f}% dos cases)",
            "threshold": "<=10%",
            "severidade": "CRITICO" if variant_ratio > 0.9 else "ALERTA",
            "impacto": f"-{13 - pts_variant_ratio(variant_ratio)} pts",
            "descricao": f"Existem {n_variants} variantes de processo distintas para {total_cases} cases. Cada case segue um caminho diferente — dificil identificar um fluxo padrao."
        })
    if top10_coverage < 0.6:
        diagnostics.append({
            "check": "Baixa concentracao nas top 10 variantes",
            "valor": f"{top10_coverage*100:.1f}%",
            "threshold": ">=80%",
            "severidade": "ALERTA",
            "impacto": f"-{17 - pts_top10(top10_coverage)} pts",
            "descricao": f"As 10 variantes mais comuns cobrem apenas {top10_coverage*100:.1f}% dos cases. O processo e muito fragmentado para gerar insights representativos."
        })
    if overlap_case_rate is not None and overlap_case_rate > 0.1:
        diagnostics.append({
            "check": "Sobreposicao de eventos no mesmo case",
            "valor": f"{overlap_case_rate*100:.1f}% dos cases",
            "threshold": "<=10%",
            "severidade": "ALERTA",
            "impacto": f"-{20 - pts_overlap(overlap_case_rate)} pts",
            "descricao": f"{overlap_case_rate*100:.1f}% dos cases tem eventos com intervalos sobrepostos. Pode ser paralelismo nao declarado ou erro de registro de timestamp."
        })
    if nr_case_p6 > 0:
        diagnostics.append({
            "check": "Nulos em Case_ID",
            "valor": f"{nr_case_p6*100:.2f}%",
            "threshold": "0%",
            "severidade": "CRITICO" if nr_case_p6 > 0.02 else "ALERTA",
            "impacto": f"-{10 - s64} pts",
            "descricao": f"{nr_case_p6*100:.2f}% dos registros estao sem Case_ID. Esses eventos ficam excluidos do calculo de variantes e cobertura de atividades."
        })

    return {
        "score": score,
        "status": status,
        "gates_triggered": gates_triggered,
        "subcategories": {
            "6.1_vocabulario_atividades": {"score": s61,                                    "max": 30},
            "6.2_variantes_processo":     {"score": s62,                                    "max": 40},
            "6.3_sobreposicao":           {"score": s63 if s63 is not None else 20,         "max": 20},  # sem end_col -> max
            "6.4_completude_case_id":     {"score": s64,                                    "max": 10},
        },
        "metrics": {
            "n_activities":         n_activities,
            "rare_activities_rate": round(rare_rate * 100, 2),
            "n_variants":           n_variants,
            "variant_ratio":        round(variant_ratio, 4),
            "top10_coverage_pct":   round(top10_coverage * 100, 2),
            "rare_variants_rate":   round(rare_variants_rate * 100, 2),
            "overlap_case_rate":    round(overlap_case_rate * 100, 2) if overlap_case_rate is not None else None,
            "null_rate_case_id":    round(nr_case_p6 * 100, 3),
        },
        "diagnostics": diagnostics,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SCORE FINAL
# ─────────────────────────────────────────────────────────────────────────────

def _barra(score, score_max=100, width=10):
    """Gera barra de progresso textual. Ex: ████████░░ 80/100"""
    filled = round((score / max(score_max, 1)) * width)
    bar    = "█" * filled + "░" * (width - filled)
    return f"{bar} {score}/{score_max}"


def _frase_pilar(pilar_nome: str, score: int, subcats: dict, gates: list) -> str:
    """
    Gera frase de diagnostico automatica para um pilar.
    Identifica o maior perdedor e monta sentenca descritiva.
    """
    SUBCAT_NAMES = {
        "1.1_presenca_colunas":       "presenca de colunas obrigatorias",
        "1.2_parse_formatos":         "qualidade dos timestamps",
        "2.1_completude_colunas":     "completude das colunas de conteudo",
        "2.2_densidade_por_case":     "densidade de eventos por case",
        "2.3_cobertura_temporal":     "cobertura temporal da base",
        "3.1_inversao_ordem":         "ordem cronologica dos eventos",
        "3.2_duracoes_negativas":     "duracoes negativas",
        "3.3_gaps_absurdos":          "gaps absurdos intra-case",
        "3.4_gap_max_eventos":        "gap maximo entre eventos",
        "4.1_duplicatas_exatas":      "duplicatas exatas",
        "4.2_colisoes":               "colisoes de registro",
        "5.1_fragmentacao":           "fragmentacao do Case ID",
        "5.2_mescla":                 "mescla de processos",
        "5.3_normalizacao_case_id":   "normalizacao do Case ID",
        "6.1_vocabulario_atividades": "vocabulario de atividades",
        "6.2_variantes_processo":     "diversidade de variantes",
        "6.3_sobreposicao":           "sobreposicao de eventos",
        "6.4_completude_case_id":     "completude do Case ID",
    }

    if gates:
        return (f"{pilar_nome} com GATE eliminatorio ativo — score {score}/100. "
                f"Corrija antes de prosseguir: {gates[0]}.")

    perdas = []
    for sk, sv in subcats.items():
        perdido = sv["max"] - sv["score"]
        if perdido > 0:
            nome = SUBCAT_NAMES.get(sk, sk)
            perdas.append((perdido, nome, sv["score"], sv["max"]))
    perdas.sort(reverse=True)

    if not perdas:
        return f"{pilar_nome} sem problemas — score {score}/100. Todos os criterios atendidos."

    # Se score == 100 mas ha subcats nao-maximas (ex: sub com teto < 100)
    # informa apenas como aviso informativo
    if score == 100:
        nomes = [p[1] for p in perdas[:2]]
        return (f"{pilar_nome} score maximo {score}/100. "
                f"Pontos nao atingidos em: {'; '.join(nomes)} — sem impacto no score final.")

    principal = perdas[0]
    frase = (f"{pilar_nome} obteve {score}/100. "
             f"Principal perda: {principal[1]} ({principal[2]} de {principal[3]} pts).")
    if len(perdas) >= 2:
        outros = [f"{p[1]} ({p[2]}/{p[3]} pts)" for p in perdas[1:3]]
        frase += f" Tambem afetado: {'; '.join(outros)}."
    return frase


def compute_final_score(pillars: dict) -> dict:
    """
    Calcula o score final ponderado pelos 6 pilares.
    Retorna score, rating, breakdown por subcategoria, ranking de foco
    e frases de diagnostico automaticas.
    """
    weights      = PILLAR_WEIGHTS
    total_weight = sum(weights.values())

    weighted_sum = sum(
        pillars[key]["score"] * weights[key]
        for key in weights if key in pillars
    )
    final_score = round(weighted_sum / total_weight, 1)

    # ── Rating ──
    if final_score >= 95:
        rating  = "SUPER_APTA"
        color   = "#1A8A4A"
        label   = "Base excelente para Process Mining"
        summary = "A base atende com folga a todos os criterios de qualidade. Excelente base para gerar visoes de Process Mining confiaveis e ricas."
    elif final_score >= 85:
        rating  = "APTA"
        color   = "#27AE60"
        label   = "Base apta para Process Mining"
        summary = "A base atende aos criterios essenciais de qualidade para gerar uma visao de Process Mining confiavel."
    elif final_score >= 75:
        rating  = "APTA_COM_RESSALVAS"
        color   = "#E67E22"
        label   = "Base apta, porem com ressalvas"
        summary = "A base e utilizavel para Process Mining, mas existem pontos de atencao que podem comprometer a qualidade dos insights."
    else:
        rating  = "NAO_APTA"
        color   = "#C0392B"
        label   = "Base nao apta — mudancas necessarias"
        summary = "A base nao atende aos criterios minimos para Process Mining. Corrija os problemas identificados antes de prosseguir."

    pilar_labels = {
        "pillar1": "P1 — Qualidade Estrutural",
        "pillar2": "P2 — Completude & Cobertura",
        "pillar3": "P3 — Integridade Temporal",
        "pillar4": "P4 — Unicidade & Duplicidade",
        "pillar5": "P5 — Coerencia do Case ID",
        "pillar6": "P6 — Minerabilidade",
    }

    # ── Breakdown por subcategoria ──
    pillar_breakdown = []
    for key in weights:
        if key not in pillars:
            continue
        pdata   = pillars[key]
        score   = pdata["score"]
        subcats = pdata.get("subcategories", {})
        gates   = pdata.get("gates_triggered", [])
        peso    = weights[key]
        nome    = pilar_labels.get(key, key)

        pts_perdidos            = 100 - score
        pts_perdidos_ponderados = round(pts_perdidos * peso / 100, 2)

        # Detalhe por subcategoria
        subs_detail = []
        for sk, sv in subcats.items():
            perdido = sv["max"] - sv["score"]
            if sv["score"] == sv["max"]:
                status_sub = "PASSOU"
            elif sv["score"] == 0:
                status_sub = "ZEROU"
            else:
                status_sub = "REDUZIU"
            subs_detail.append({
                "subcategoria":  sk,
                "score_obtido":  sv["score"],
                "score_max":     sv["max"],
                "pts_perdidos":  perdido,
                "status":        status_sub,
                "barra":         _barra(sv["score"], sv["max"]),
                "impacto_texto": f"obteve {sv['score']} de {sv['max']} (-{perdido} pts)" if perdido > 0 else f"maximo atingido ({sv['score']}/{sv['max']})",
            })
        subs_detail.sort(key=lambda x: x["pts_perdidos"], reverse=True)

        pillar_breakdown.append({
            "pilar_key":               key,
            "pilar":                   nome,
            "score":                   score,
            "barra_pilar":             _barra(score),
            "pts_perdidos":            pts_perdidos,
            "pts_perdidos_ponderados": pts_perdidos_ponderados,
            "peso":                    peso,
            "status":                  pdata.get("status", "APTO"),
            "gates":                   gates,
            "subcategorias":           subs_detail,
            "frase_diagnostico":       _frase_pilar(nome, score, subcats, gates),
        })

    # ── Ranking de foco (maior impacto ponderado primeiro) ──
    pillar_breakdown.sort(key=lambda x: x["pts_perdidos_ponderados"], reverse=True)
    ranking_str = " > ".join(
        p["pilar_key"].replace("pillar", "P")
        for p in pillar_breakdown
        if p["pts_perdidos"] > 0
    ) or "Nenhum problema"

    # ── Diagnosticos consolidados (compatibilidade) ──
    all_diagnostics = []
    sev_order = {"CRITICO": 0, "ALERTA": 1, "WARN_FORTE": 2, "WARN": 3, "INFO": 4}
    for key, pdata in pillars.items():
        for d in pdata.get("diagnostics", []):
            all_diagnostics.append({**d, "pilar": pilar_labels.get(key, key)})
    all_diagnostics.sort(key=lambda x: sev_order.get(x.get("severidade", "INFO"), 5))

    # Gates consolidados
    all_gates = []
    for key, pdata in pillars.items():
        for g in pdata.get("gates_triggered", []):
            all_gates.append({"pilar": pilar_labels.get(key, key), "gate": g})

    # Frase resumo global
    frases_problema = [
        p["frase_diagnostico"]
        for p in pillar_breakdown
        if p["pts_perdidos"] > 0
    ]
    frase_global = " | ".join(frases_problema[:3]) if frases_problema else "Base sem problemas identificados."

    return {
        "final_score":       final_score,
        "barra_final":       _barra(final_score),
        "rating":            rating,
        "color":             color,
        "label":             label,
        "summary":           summary,
        "ranking_foco":      ranking_str,
        "frase_resumo":      frase_global,
        "gates_triggered":   all_gates,
        "diagnostics":       all_diagnostics,
        "pillar_breakdown":  pillar_breakdown,
        "pillar_scores":     {k: pillars[k]["score"] for k in weights if k in pillars},
        "pillar_status":     {k: pillars[k]["status"] for k in weights if k in pillars},
    }


# ─────────────────────────────────────────────────────────────────────────────
# MAIN — PIPELINE COMPLETO
# ─────────────────────────────────────────────────────────────────────────────

def run_analysis(input_path: str, case_col: str, act_col: str,
                 start_col: str, end_col=None,
                 delimiter: str = "auto", encoding: str = "auto",
                 dayfirst: bool = None,
                 original_filename: str = None) -> dict:
    """
    Executa os 6 pilares e retorna o resultado completo como dict.
    Inclui metadados de tempo, tamanho da base e risco consultivo.

    dayfirst=None (padrao): inferido automaticamente pelo formato detectado.
    original_filename: nome original do arquivo (usado no log em vez do path temp).
    """
    ts_inicio = datetime.now()

    # ── H: Cache — tenta carregar DataFrame pré-processado ──
    _cache = _get_cache() if _CACHE_ENABLED else None
    _cache_key = input_path  # chave = path do arquivo (mtime invalida automaticamente)
    df = None
    df_sorted = None
    dayfirst_used = None
    ts_inferred = None

    if _cache:
        _cached = _cache.load(_cache_key)
        if _cached is not None:
            df        = _cached.get("df")
            df_sorted = _cached.get("df_sorted")
            dayfirst_used = _cached.get("dayfirst_used")
            ts_inferred   = _cached.get("ts_inferred")

    if df is None:
        # Cache miss — carrega e processa do zero
        df_raw = load_data(input_path, delimiter, encoding)
        validate_columns(df_raw, case_col, act_col, start_col, end_col)
        df = normalize_strings(df_raw)

        # ── Inferir formato e dayfirst automaticamente ──
        ts_series = df[start_col].dropna()
        if end_col and end_col in df.columns:
            ts_series = pd.concat([ts_series, df[end_col].dropna()])
        ts_inferred = infer_datetime_format(ts_series)

        if dayfirst is None:
            dayfirst_used = ts_inferred["dayfirst"]
        else:
            dayfirst_used = dayfirst

        # ── B + F + G: timestamps, category dtype, sort único ──
        df, df_sorted = prepare_df(df, case_col, act_col, start_col, end_col, dayfirst_used)

        # Salvar no cache para próximas análises
        if _cache:
            _cache.save(_cache_key, {
                "df": df,
                "df_sorted": df_sorted,
                "dayfirst_used": dayfirst_used,
                "ts_inferred": ts_inferred,
            })
    else:
        # Cache hit — dayfirst pode ter sido sobrescrito pelo chamador
        if dayfirst is not None and dayfirst != dayfirst_used:
            # Chamador passou dayfirst explícito diferente: reprocessar timestamps
            dayfirst_used = dayfirst
            df["__start_ts"] = pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst_used)
            if end_col:
                df["__end_ts"] = pd.to_datetime(df[end_col], errors="coerce", dayfirst=dayfirst_used)
            df_sorted = df.sort_values([case_col, "__start_ts"]).reset_index(drop=True)

    pillar_results = {
        "pillar1": score_pillar1(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used, ts_inferred),
        "pillar2": score_pillar2(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used),
        "pillar3": score_pillar3(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used, df_sorted),
        "pillar4": score_pillar4(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used),
        "pillar5": score_pillar5(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used, df_sorted),
        "pillar6": score_pillar6(df, case_col, act_col, start_col, end_col,
                                 dayfirst_used, df_sorted),
    }

    final = compute_final_score(pillar_results)

    ts_fim = datetime.now()
    tempo_minutos = round((ts_fim - ts_inicio).total_seconds() / 60, 4)

    # ── Metadados da base ──
    n_cases      = int(df[case_col].nunique())
    n_atividades = int(df[act_col].nunique())

    try:
        # B: usa __start_ts já parseado
        _ts = df["__start_ts"] if "__start_ts" in df.columns else pd.to_datetime(df[start_col], errors="coerce", dayfirst=dayfirst_used)
        periodo_dias = int((_ts.max() - _ts.min()).days)
    except Exception:
        periodo_dias = 0

    # ── Risco consultivo (media de P3 a P6) ──
    scores_consultivos = [
        pillar_results["pillar3"]["score"],
        pillar_results["pillar4"]["score"],
        pillar_results["pillar5"]["score"],
        pillar_results["pillar6"]["score"],
    ]
    media_consultiva = round(sum(scores_consultivos) / len(scores_consultivos), 1)
    if media_consultiva >= 80:
        risco_consultivo = "BAIXO"
    elif media_consultiva >= 60:
        risco_consultivo = "MEDIO"
    else:
        risco_consultivo = "ALTO"

    # ── Veredicto eliminatorio (baseado em P1 + P2 + gates) ──
    p1_ok = pillar_results["pillar1"]["status"] == "APTO"
    p2_ok = pillar_results["pillar2"]["status"] == "APTO"
    veredicto = "APTA" if (p1_ok and p2_ok) else "NAO_APTA"

    # Nome do arquivo: usa original_filename se fornecido, senao basename do path
    display_name = original_filename if original_filename else os.path.basename(input_path)

    return {
        "input_file":   display_name,
        "n_rows":       len(df),
        "n_cols":       len(df.columns),
        "n_cases":      n_cases,
        "n_atividades": n_atividades,
        "periodo_dias": periodo_dias,
        "ts_inicio":    ts_inicio.strftime("%Y-%m-%d %H:%M:%S"),
        "ts_fim":       ts_fim.strftime("%Y-%m-%d %H:%M:%S"),
        "tempo_minutos": tempo_minutos,
        "veredicto":    veredicto,
        "risco_consultivo": risco_consultivo,
        "media_consultiva": media_consultiva,
        "ts_format_inferred": {
            "formato":    ts_inferred["fmt"],
            "dayfirst":   dayfirst_used,
            "n_formatos": ts_inferred["n_formats"],
            "parse_rate": ts_inferred["parse_rate"],
            "label":      ts_inferred["label"],
        },
        "column_mapping": {
            "case_id":          case_col,
            "activity":         act_col,
            "timestamp_inicio": start_col,
            "timestamp_fim":    end_col,
        },
        "final": final,
        "pillars": pillar_results,
    }


# ─────────────────────────────────────────────────────────────────────────────
# LOG DE SAÍDA — TABELAS execucoes + diagnosticos
# ─────────────────────────────────────────────────────────────────────────────

def _get_responsavel() -> str:
    """Retorna o usuário do sistema operacional que executou a análise."""
    try:
        return getpass.getuser()
    except Exception:
        return "desconhecido"


def _build_log_rows(result: dict, id_execucao: int) -> tuple:
    """
    A partir do resultado de run_analysis, monta:
      - row_exec: dict com 1 linha para a tabela execucoes
      - rows_diag: lista de dicts para a tabela diagnosticos
    """
    final    = result["final"]
    pillars  = result["pillars"]

    # ── Campos derivados para indicadores ──────────────────────────────────────
    # AnoMes da execucao: YYYYMM (ex: 202604)
    anomes = result["ts_inicio"][:7].replace("-", "")   # "2026-04-..." -> "202604"

    # Tempo formatado como MM:SS,ms (ex: 01:23,456)
    tempo_seg_total = result["tempo_minutos"] * 60       # converte para segundos
    mm  = int(tempo_seg_total // 60)
    ss  = int(tempo_seg_total % 60)
    ms  = int(round((tempo_seg_total - int(tempo_seg_total)) * 1000))
    tempo_fmt = f"{mm:02d}:{ss:02d},{ms:03d}"

    # Soma dos scores determinativos (P1 + P2) — veredicto
    soma_p1_p2 = round(pillars["pillar1"]["score"] + pillars["pillar2"]["score"], 1)

    # Soma dos scores consultivos (P3 + P4 + P5 + P6)
    soma_p3_p6 = round(
        pillars["pillar3"]["score"] + pillars["pillar4"]["score"] +
        pillars["pillar5"]["score"] + pillars["pillar6"]["score"], 1
    )

    # ── Linha de execução ──
    row_exec = {
        "id_execucao":       id_execucao,
        "anomes":            int(anomes),
        "nome_base":         result["input_file"],
        "responsavel":       _get_responsavel(),
        "data_execucao":     result["ts_inicio"][:10],
        "ts_inicio":         result["ts_inicio"],
        "ts_fim":            result["ts_fim"],
        "tempo_execucao":    tempo_fmt,
        # Tamanho
        "n_linhas":          result["n_rows"],
        "n_cases":           result["n_cases"],
        "n_atividades":      result["n_atividades"],
        "periodo_dias":      result["periodo_dias"],
        # Veredicto (P1 + P2)
        "score_p1":          pillars["pillar1"]["score"],
        "score_p2":          pillars["pillar2"]["score"],
        "soma_p1_p2":        soma_p1_p2,
        "veredicto":         result["veredicto"],
        "n_gates_disparados": len(final["gates_triggered"]),
        # Consultivos (P3–P6)
        "score_p3":          pillars["pillar3"]["score"],
        "score_p4":          pillars["pillar4"]["score"],
        "score_p5":          pillars["pillar5"]["score"],
        "score_p6":          pillars["pillar6"]["score"],
        "soma_p3_p6":        soma_p3_p6,
        "media_consultiva":  result["media_consultiva"],
        "risco_consultivo":  result["risco_consultivo"],
        # Resumo geral
        "score_final":       final["final_score"],
        "barra_final":       final.get("barra_final", ""),
        "classificacao":     final["rating"],
        "n_diagnosticos_total": len(final["diagnostics"]),
        # Onde focar e resumo narrativo
        "ranking_foco":      final.get("ranking_foco", ""),
        "frase_resumo":      final.get("frase_resumo", ""),
    }

    # ── Linhas de subcategorias (1 por sub por execucao) ──
    rows_sub = []
    seq_sub  = 1
    for pb in final.get("pillar_breakdown", []):
        pilar_key   = pb["pilar_key"]
        pilar_label = pb["pilar_key"].replace("pillar", "P")
        for sd in pb.get("subcategorias", []):
            rows_sub.append({
                "id_execucao":   id_execucao,
                "seq":           seq_sub,
                "nome_base":     result["input_file"],
                "pilar":         pilar_label,
                "subcategoria":  sd["subcategoria"],
                "score_obtido":  sd["score_obtido"],
                "score_max":     sd["score_max"],
                "pts_perdidos":  sd["pts_perdidos"],
                "status":        sd["status"],
                "barra":         sd["barra"],
                "impacto_texto": sd["impacto_texto"],
            })
            seq_sub += 1

    # ── Linhas de diagnóstico ──
    # Mapa de chave interna → rótulo curto do pilar
    pilar_map = {
        "pillar1": "P1", "pillar2": "P2", "pillar3": "P3",
        "pillar4": "P4", "pillar5": "P5", "pillar6": "P6",
    }
    # Tipo por severidade
    tipo_map = {
        "CRITICO":    "GATE",
        "ALERTA":     "AVISO",
        "WARN_FORTE": "AVISO",
        "WARN":       "AVISO",
        "INFO":       "INFO",
    }
    # Impacto por severidade
    impacto_map = {
        "CRITICO":    "ELIMINATORIO",
        "ALERTA":     "REDUZ_SCORE",
        "WARN_FORTE": "REDUZ_SCORE",
        "WARN":       "REDUZ_SCORE",
        "INFO":       "INFORMATIVO",
    }

    # Mapa: palavras-chave do check → subcategoria por pilar
    # Cada entrada: (pilar_key, palavras_no_check) -> subcat_key
    # Ordem importa: mais específico primeiro
    SUBCAT_KEYWORDS = {
        "pillar1": [
            (["nulos em case_id", "case_id nulo"],        "1.1_presenca_colunas"),
            (["nulos em atividade", "atividade nula"],     "1.1_presenca_colunas"),
            (["timestamp_inicio", "timestamp_fim"],        "1.1_presenca_colunas"),
            (["timestamp_fim com", "fim invalido"],        "1.1_presenca_colunas"),
            (["multiplos formatos", "formato"],            "1.2_parse_formatos"),
            (["invalido", "parse"],                        "1.2_parse_formatos"),
        ],
        "pillar2": [
            (["nulos em atividade"],                       "2.1_completude_colunas"),
            (["nulos em timestamp"],                       "2.1_completude_colunas"),
            (["cases com 1 evento", "pct_lt2", "mediana"], "2.2_densidade_por_case"),
            (["periodo", "cobertura", "temporal"],         "2.3_cobertura_temporal"),
        ],
        "pillar3": [
            (["inversao", "inversão", "ordem"],            "3.1_inversao_ordem"),
            (["duracao negativa", "duração negativa",
              "negativas"],                                "3.2_duracoes_negativas"),
            (["gap absurdo", "absurdo"],                   "3.3_gaps_absurdos"),
            (["gap sem eventos", "gap max", "gap máximo"], "3.4_gap_max_eventos"),
        ],
        "pillar4": [
            (["duplicata", "duplicatas exatas"],           "4.1_duplicatas_exatas"),
            (["colisao", "colisão", "ambiguidade"],        "4.2_colisoes"),
        ],
        "pillar5": [
            (["fragmentacao", "fragmentação", "r_duration",
              "duracao do case"],                          "5.1_fragmentacao"),
            (["mescla", "r_events", "monster"],            "5.2_mescla"),
            (["case_id inconsistente", "normalizacao",
              "normalização", "espacos", "controle"],      "5.3_normalizacao_case_id"),
        ],
        "pillar6": [
            (["atividades raras", "cardinali"],            "6.1_vocabulario_atividades"),
            (["variante", "variant", "top10"],             "6.2_variantes_processo"),
            (["sobreposicao", "sobreposição", "overlap"],  "6.3_sobreposicao"),
            (["nulos em case_id", "case_id nulo"],         "6.4_completude_case_id"),
        ],
    }

    def _find_subcat(pilar_key, check_str, subcats):
        """Encontra a subcategoria correta para um diagnóstico via check."""
        check_lower = check_str.lower()
        for keywords, subcat_key in SUBCAT_KEYWORDS.get(pilar_key, []):
            if any(kw in check_lower for kw in keywords):
                if subcat_key in subcats:
                    return subcat_key
        # Fallback: primeira subcategoria do pilar
        return list(subcats.keys())[0] if subcats else ""

    def _fmt_impacto(score_obtido, score_max):
        """Formata impacto como 'obteve X de Y (-Z pts)'."""
        perdido = score_max - score_obtido
        return f"obteve {score_obtido} de {score_max} (-{perdido} pts)"

    rows_diag = []
    seq = 1
    for pilar_key, pdata in pillars.items():
        pilar_label = pilar_map.get(pilar_key, pilar_key)
        subcats     = pdata.get("subcategories", {})

        for diag in pdata.get("diagnostics", []):
            sev   = diag.get("severidade", "INFO")
            check = diag.get("check", "")

            # Encontrar subcategoria correspondente
            subcat_key = _find_subcat(pilar_key, check, subcats)
            subcat_data = subcats.get(subcat_key, {})
            score_obtido = subcat_data.get("score", 0)
            score_max    = subcat_data.get("max",   0)

            rows_diag.append({
                "id_execucao":       id_execucao,
                "seq":               seq,
                "nome_base":         result["input_file"],
                "pilar":             pilar_label,
                "subcategoria":      subcat_key,
                "score_obtido":      score_obtido,
                "score_max":         score_max,
                "tipo":              tipo_map.get(sev, "INFO"),
                "check":             check,
                "mensagem":          diag.get("descricao", ""),
                "valor_encontrado":  diag.get("valor", ""),
                "limite_referencia": diag.get("threshold", ""),
                "impacto_pts":       _fmt_impacto(score_obtido, score_max),
                "impacto":           impacto_map.get(sev, "INFORMATIVO"),
                "severidade":        sev,
            })
            seq += 1

        # Gates sem diagnóstico explícito associado
        for gate_msg in pdata.get("gates_triggered", []):
            already = any(
                gate_msg.lower()[:20] in d.get("mensagem", "").lower()
                for d in rows_diag
                if d["id_execucao"] == id_execucao and d["pilar"] == pilar_label
            )
            if not already:
                rows_diag.append({
                    "id_execucao":       id_execucao,
                    "seq":               seq,
                    "nome_base":         result["input_file"],
                    "pilar":             pilar_label,
                    "subcategoria":      "",
                    "score_obtido":      0,
                    "score_max":         0,
                    "tipo":              "GATE",
                    "check":             "Gate eliminatorio",
                    "mensagem":          gate_msg,
                    "valor_encontrado":  "",
                    "limite_referencia": "",
                    "impacto_pts":       "score zerado (gate)",
                    "impacto":           "ELIMINATORIO",
                    "severidade":        "CRITICO",
                })
                seq += 1

    return row_exec, rows_diag, rows_sub


def export_to_excel(result: dict, excel_path: str = "pm_engine_log.xlsx") -> str:
    """
    Grava ou acrescenta os resultados de uma execução em um arquivo Excel.

    Estrutura:
      - Aba 'execucoes'   : 1 linha por execução (Tabela 1)
      - Aba 'diagnosticos': N linhas por execução (Tabela 2)

    Se o arquivo não existir, cria do zero.
    Se já existir, acrescenta linhas mantendo histórico.

    Retorna o caminho do arquivo salvo.
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instale openpyxl: pip install openpyxl")

    # ── Determinar próximo id_execucao ──
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws_exec = wb["execucoes"]   if "execucoes"    in wb.sheetnames else wb.create_sheet("execucoes")
        ws_diag = wb["diagnosticos"] if "diagnosticos" in wb.sheetnames else wb.create_sheet("diagnosticos")

        # Próximo ID = max atual + 1
        ids_existentes = [
            ws_exec.cell(row=r, column=1).value
            for r in range(2, ws_exec.max_row + 1)
            if ws_exec.cell(row=r, column=1).value is not None
        ]
        id_execucao = (max(int(x) for x in ids_existentes if str(x).isdigit()) + 1) if ids_existentes else 1
    else:
        wb       = openpyxl.Workbook()
        ws_exec  = wb.active
        ws_exec.title = "execucoes"
        ws_diag  = wb.create_sheet("diagnosticos")
        id_execucao = 1

    # ── Montar linhas ──
    row_exec, rows_diag, rows_sub = _build_log_rows(result, id_execucao)

    def _next_data_row(ws) -> int:
        """
        Retorna o índice da próxima linha disponível para escrita,
        ignorando linhas que têm formatação mas nenhum valor.
        Percorre a coluna 1 procurando o último valor real e retorna
        a linha seguinte — garante append linear sem buracos.
        """
        last_filled = 1  # linha do cabeçalho
        for r in range(2, ws.max_row + 2):
            if ws.cell(row=r, column=1).value is not None:
                last_filled = r
        return last_filled + 1

    def _write_row(ws, row_idx, values):
        """Escreve uma lista de valores em células consecutivas a partir da col 1."""
        for c, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=c, value=v)

    # ── Estilos ──
    BRAND     = "1F4E79"
    ACCENT    = "2E74B5"
    LIGHT     = "D6E4F0"
    ALT       = "EBF3FB"
    WHITE     = "FFFFFF"
    RED_FILL  = "FADBD8"
    ORG_FILL  = "FDEBD0"
    GRN_FILL  = "D5F5E3"

    hdr_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    hdr_fill  = PatternFill("solid", fgColor=BRAND)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(border_style="thin", color="C0C8D0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)

    def style_header(ws, headers):
        # Escrever diretamente na linha 1 (evita inserção em linha errada)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
        ws.row_dimensions[1].height = 32

    def style_data_row(ws, row_idx, n_cols, alt=False):
        fill_color = ALT if alt else WHITE
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = cell_border
            cell.fill      = PatternFill("solid", fgColor=fill_color)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── ABA execucoes ──
    EXEC_HEADERS = [
        "id_execucao", "anomes", "nome_base", "responsavel", "data_execucao", "ts_inicio", "ts_fim",
        "tempo_execucao",
        "n_linhas", "n_cases", "n_atividades", "periodo_dias",
        "score_p1", "score_p2", "soma_p1_p2", "veredicto", "n_gates_disparados",
        "score_p3", "score_p4", "score_p5", "score_p6", "soma_p3_p6",
        "media_consultiva", "risco_consultivo",
        "score_final", "barra_final", "classificacao", "n_diagnosticos_total",
        "ranking_foco", "frase_resumo",
    ]

    # Inserir cabecalho apenas se a celula A1 ainda nao tiver o header
    if ws_exec.cell(1, 1).value != "id_execucao":
        style_header(ws_exec, EXEC_HEADERS)
        set_col_widths(ws_exec, [
            6, 10, 28, 20, 14, 20, 20,
            14,
            10, 10, 12, 12,
            10, 10, 12, 14, 14,
            10, 10, 10, 10, 12,
            16, 16,
            12, 18, 22, 14,
            28, 80,
        ])
        ws_exec.freeze_panes = "A2"

    exec_row_data = [row_exec[h] for h in EXEC_HEADERS]
    new_row_idx = _next_data_row(ws_exec)
    _write_row(ws_exec, new_row_idx, exec_row_data)
    alt = (new_row_idx % 2 == 0)
    style_data_row(ws_exec, new_row_idx, len(EXEC_HEADERS), alt)

    # ── Coloração semântica das células ──────────────────────────────────────
    col_veredicto = EXEC_HEADERS.index("veredicto") + 1
    col_classif   = EXEC_HEADERS.index("classificacao") + 1
    col_risco     = EXEC_HEADERS.index("risco_consultivo") + 1
    col_soma12    = EXEC_HEADERS.index("soma_p1_p2") + 1
    col_soma36    = EXEC_HEADERS.index("soma_p3_p6") + 1
    col_anomes    = EXEC_HEADERS.index("anomes") + 1
    col_tempo     = EXEC_HEADERS.index("tempo_execucao") + 1

    # Veredicto
    v_cell = ws_exec.cell(new_row_idx, col_veredicto)
    if v_cell.value == "APTA":
        v_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
        v_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)
    else:
        v_cell.fill = PatternFill("solid", fgColor=RED_FILL)
        v_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)

    # Classificacao
    c_cell = ws_exec.cell(new_row_idx, col_classif)
    if "NAO" in str(c_cell.value):
        c_cell.fill = PatternFill("solid", fgColor=RED_FILL)
        c_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
    elif "RESSALVAS" in str(c_cell.value):
        c_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
        c_cell.font = Font(name="Arial", size=10, color="784212", bold=True)
    elif "SUPER" in str(c_cell.value):
        c_cell.fill = PatternFill("solid", fgColor="A9DFBF")   # verde mais escuro
        c_cell.font = Font(name="Arial", size=10, color="0B5226", bold=True)
    else:
        c_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
        c_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)

    # Risco consultivo
    r_cell = ws_exec.cell(new_row_idx, col_risco)
    if r_cell.value == "ALTO":
        r_cell.fill = PatternFill("solid", fgColor=RED_FILL)
        r_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
    elif r_cell.value == "MEDIO":
        r_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
        r_cell.font = Font(name="Arial", size=10, color="784212", bold=True)
    else:
        r_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
        r_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)

    # AnoMes — destaque azul suave para facilitar filtro/agrupamento
    am_cell = ws_exec.cell(new_row_idx, col_anomes)
    am_cell.fill = PatternFill("solid", fgColor="D6E4F0")
    am_cell.font = Font(name="Arial", size=10, color="1F4E79", bold=True)
    am_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Tempo de execucao — fonte monoespaçada para legibilidade MM:SS,ms
    t_cell = ws_exec.cell(new_row_idx, col_tempo)
    t_cell.font = Font(name="Courier New", size=10, color="555555")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Soma P1+P2 — coloração por faixa (indicador determinativo)
    s12_cell = ws_exec.cell(new_row_idx, col_soma12)
    s12_val  = s12_cell.value or 0
    if s12_val >= 180:   # >= 90 cada = excelente
        s12_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
        s12_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)
    elif s12_val >= 140:
        s12_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
        s12_cell.font = Font(name="Arial", size=10, color="784212", bold=True)
    else:
        s12_cell.fill = PatternFill("solid", fgColor=RED_FILL)
        s12_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
    s12_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Soma P3+P6 — coloração por faixa (indicador consultivo)
    s36_cell = ws_exec.cell(new_row_idx, col_soma36)
    s36_val  = s36_cell.value or 0
    if s36_val >= 320:   # >= 80 em média nos 4
        s36_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
        s36_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)
    elif s36_val >= 240:
        s36_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
        s36_cell.font = Font(name="Arial", size=10, color="784212", bold=True)
    else:
        s36_cell.fill = PatternFill("solid", fgColor=RED_FILL)
        s36_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
    s36_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── ABA diagnosticos ──
    DIAG_HEADERS = [
        "id_execucao", "seq", "nome_base", "pilar", "subcategoria",
        "score_obtido", "score_max", "impacto_pts",
        "tipo", "check", "mensagem",
        "valor_encontrado", "limite_referencia",
        "impacto", "severidade",
    ]

    if ws_diag.cell(1, 1).value != "id_execucao":
        style_header(ws_diag, DIAG_HEADERS)
        set_col_widths(ws_diag, [
            10, 6, 24, 6, 26,
            12, 10, 26,
            8, 30, 65,
            18, 18,
            14, 12,
        ])
        ws_diag.freeze_panes = "A2"

    col_tipo_diag = DIAG_HEADERS.index("tipo") + 1
    col_imp_diag  = DIAG_HEADERS.index("impacto") + 1
    col_sev_diag  = DIAG_HEADERS.index("severidade") + 1

    for i, drow in enumerate(rows_diag):
        diag_row_data = [drow[h] for h in DIAG_HEADERS]
        diag_idx = _next_data_row(ws_diag)
        _write_row(ws_diag, diag_idx, diag_row_data)
        alt_d = (diag_idx % 2 == 0)
        style_data_row(ws_diag, diag_idx, len(DIAG_HEADERS), alt_d)

        # Colorir tipo e severidade
        t_cell = ws_diag.cell(diag_idx, col_tipo_diag)
        if t_cell.value == "GATE":
            t_cell.fill = PatternFill("solid", fgColor=RED_FILL)
            t_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
        elif t_cell.value == "AVISO":
            t_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
            t_cell.font = Font(name="Arial", size=10, color="784212", bold=True)

        s_cell = ws_diag.cell(diag_idx, col_sev_diag)
        if s_cell.value == "CRITICO":
            s_cell.fill = PatternFill("solid", fgColor=RED_FILL)
            s_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
        elif s_cell.value in ("ALERTA", "WARN_FORTE"):
            s_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
            s_cell.font = Font(name="Arial", size=10, color="784212", bold=True)

    # ── ABA subcategorias ──
    SUB_HEADERS = [
        "id_execucao", "seq", "nome_base", "pilar", "subcategoria",
        "score_obtido", "score_max", "pts_perdidos",
        "status", "barra", "impacto_texto",
    ]

    ws_sub = wb["subcategorias"] if "subcategorias" in wb.sheetnames else wb.create_sheet("subcategorias")

    if ws_sub.cell(1, 1).value != "id_execucao":
        style_header(ws_sub, SUB_HEADERS)
        set_col_widths(ws_sub, [10, 6, 24, 6, 32, 12, 10, 12, 10, 18, 34])
        ws_sub.freeze_panes = "A2"

    col_status_sub = SUB_HEADERS.index("status") + 1

    for srow in rows_sub:
        sub_row_data = [srow[h] for h in SUB_HEADERS]
        sub_idx = _next_data_row(ws_sub)
        _write_row(ws_sub, sub_idx, sub_row_data)
        alt_s   = (sub_idx % 2 == 0)
        style_data_row(ws_sub, sub_idx, len(SUB_HEADERS), alt_s)

        # Colorir status
        st_cell = ws_sub.cell(sub_idx, col_status_sub)
        if st_cell.value == "ZEROU":
            st_cell.fill = PatternFill("solid", fgColor=RED_FILL)
            st_cell.font = Font(name="Arial", size=10, color="922B21", bold=True)
        elif st_cell.value == "REDUZIU":
            st_cell.fill = PatternFill("solid", fgColor=ORG_FILL)
            st_cell.font = Font(name="Arial", size=10, color="784212", bold=True)
        elif st_cell.value == "PASSOU":
            st_cell.fill = PatternFill("solid", fgColor=GRN_FILL)
            st_cell.font = Font(name="Arial", size=10, color="1D6A39", bold=True)

    wb.save(excel_path)
    return os.path.abspath(excel_path)


# ─────────────────────────────────────────────────────────────────────────────
# SERVIDOR FLASK (integração com HTML)
# ─────────────────────────────────────────────────────────────────────────────

def start_server(port: int = 5000):
    """Inicia servidor Flask para integração com o frontend HTML."""
    try:
        from flask import Flask, request, jsonify
        from flask_cors import CORS
        import tempfile
    except ImportError:
        print("Instale: pip install flask flask-cors")
        sys.exit(1)

    app = Flask(__name__)
    CORS(app)
    app.config['MAX_CONTENT_LENGTH'] = 1_610_612_736  # 1.5 GB

    @app.route("/health", methods=["GET"])
    def health():
        _cache_info = {}
        if _CACHE_ENABLED:
            try:
                _cache_info = _get_cache().info()
            except Exception:
                pass
        return jsonify({"status": "ok", "engine": "pm_engine v1.0", "cache": _cache_info})

    @app.route("/cache/clear", methods=["POST"])
    def cache_clear():
        if not _CACHE_ENABLED:
            return jsonify({"cleared": 0, "message": "cache desabilitado"})
        try:
            n = _get_cache().clear()
            return jsonify({"cleared": n})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/columns", methods=["POST"])
    def columns():
        """
        Lê apenas o cabeçalho e conta linhas via bytes — muito mais rápido
        que carregar o DataFrame completo só para listar colunas.
        """
        try:
            if "file" not in request.files:
                return jsonify({"error": "Nenhum arquivo enviado"}), 400

            file = request.files["file"]
            config = json.loads(request.form.get("config", "{}"))
            suffix = Path(file.filename).suffix.lower()

            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name

            try:
                if suffix in (".xlsx", ".xls", ".ods"):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                        ws = wb.active
                        headers = []
                        n_rows = 0
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i == 0:
                                headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                            else:
                                n_rows += 1
                        wb.close()
                    except Exception:
                        df_h = pd.read_excel(tmp_path, dtype=str, nrows=0)
                        headers = list(df_h.columns)
                        df_full = pd.read_excel(tmp_path, dtype=str)
                        n_rows = len(df_full)
                    return jsonify({
                        "columns": headers,
                        "n_rows": n_rows,
                        "encoding_used": "xlsx",
                        "delimiter_detected": None,
                    })
                else:
                    detected_enc = _detect_encoding(tmp_path)
                    # Bug 1 fix: /columns always auto-detects the delimiter.
                    # The UI dropdown default (',') must not override detection here —
                    # the correct sep is returned to the frontend which then updates the UI.
                    if suffix == ".tsv":
                        sep = "\t"
                    else:
                        sep = _detect_delimiter(tmp_path, detected_enc)

                    try:
                        df_h = pd.read_csv(tmp_path, dtype=str, sep=sep,
                                           encoding=detected_enc, nrows=0)
                    except Exception:
                        df_h = pd.read_csv(tmp_path, dtype=str, sep=sep,
                                           encoding="latin-1", nrows=0)
                    headers = list(df_h.columns)

                    # Conta linhas via leitura de bytes (evita pandas para arquivo inteiro)
                    n_rows = 0
                    try:
                        with open(tmp_path, "rb") as fbin:
                            buf = bytearray(1024 * 1024)
                            while True:
                                nb = fbin.readinto(buf)
                                if not nb:
                                    break
                                n_rows += buf[:nb].count(b"\n")
                        n_rows = max(0, n_rows - 1)  # desconta header
                    except Exception:
                        n_rows = -1

                    return jsonify({
                        "columns": headers,
                        "n_rows": n_rows,
                        "encoding_used": detected_enc,
                        "delimiter_detected": sep,
                    })
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/analyze", methods=["POST"])
    def analyze():
        try:
            if "file" not in request.files:
                return jsonify({"error": "Nenhum arquivo enviado"}), 400

            file    = request.files["file"]
            mapping = json.loads(request.form.get("mapping", "{}"))
            config  = json.loads(request.form.get("config",  "{}"))

            case_col  = mapping.get("case_id")
            act_col   = mapping.get("activity")
            start_col = mapping.get("timestamp_inicio")
            # Normaliza end_col: "null", "", None -> None
            _end_raw  = mapping.get("timestamp_fim")
            end_col   = None if not _end_raw or str(_end_raw).lower() in ("null","none","") else _end_raw

            if not all([case_col, act_col, start_col]):
                return jsonify({"error": "Mapeamento incompleto: case_id, activity e timestamp_inicio sao obrigatorios"}), 400

            suffix = Path(file.filename).suffix.lower()
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name

            try:
                # Preservar nome original antes de qualquer processamento
                original_name = file.filename or os.path.basename(tmp_path)

                result = run_analysis(
                    input_path=tmp_path,
                    case_col=case_col,
                    act_col=act_col,
                    start_col=start_col,
                    end_col=end_col,
                    delimiter=config.get("delimiter", "auto"),
                    encoding=config.get("encoding", "auto"),
                    dayfirst=None,  # inferido automaticamente
                    original_filename=original_name,
                )

                # Garantir que input_file sempre reflete o nome original
                result["input_file"] = original_name

                # Gravar log no Excel — caminho absoluto relativo ao servidor
                excel_path = config.get("excel_log", "pm_engine_log.xlsx")
                excel_path = os.path.abspath(excel_path)
                try:
                    saved_path = export_to_excel(result, excel_path)
                    result["excel_log"] = saved_path
                except Exception as e_xls:
                    result["excel_log_error"] = str(e_xls)
                return jsonify(result)
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

        except ValueError as e:
            return jsonify({"error": str(e)}), 422
        except Exception as e:
            return jsonify({"error": f"Erro interno: {str(e)}"}), 500

    print(f"\n  PM Engine rodando em http://localhost:{port}")
    print(f"  Endpoints: GET /health  |  POST /analyze\n")
    app.run(host="0.0.0.0", port=port, debug=False)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PM Engine — Validador de bases para Process Mining")
    parser.add_argument("--serve",     action="store_true", help="Inicia servidor Flask")
    parser.add_argument("--port",      type=int, default=5000)
    parser.add_argument("--input",     help="Caminho do arquivo CSV/XLSX")
    parser.add_argument("--case_id",   help="Nome da coluna Case ID")
    parser.add_argument("--activity",  help="Nome da coluna Atividade")
    parser.add_argument("--start_ts",  help="Nome da coluna Timestamp Inicio")
    parser.add_argument("--end_ts",    help="Nome da coluna Timestamp Fim (opcional)")
    parser.add_argument("--delimiter", default="auto")
    parser.add_argument("--encoding",  default="utf-8")
    parser.add_argument("--dayfirst",  action="store_true", default=True)
    parser.add_argument("--output",    help="Caminho do JSON de saida (opcional)")
    args = parser.parse_args()

    if args.serve:
        start_server(args.port)
    else:
        if not all([args.input, args.case_id, args.activity, args.start_ts]):
            parser.print_help()
            sys.exit(1)

        result = run_analysis(
            input_path=args.input,
            case_col=args.case_id,
            act_col=args.activity,
            start_col=args.start_ts,
            end_col=args.end_ts,
            delimiter=args.delimiter,
            encoding=args.encoding,
            dayfirst=args.dayfirst,
        )

        output_str = json.dumps(result, ensure_ascii=False, indent=2, default=str)
        print(output_str)

        if args.output:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(output_str)
            print(f"\n  Resultado salvo em: {args.output}", file=sys.stderr)
